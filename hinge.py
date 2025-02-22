from flo import DGraph, DParams, DNode, DEdge, to_kelvin
from named_types import FloParamsHouse0, PriceQuantityUnitless
from fake_models import MessageSql
import json
from typing import List
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.colors import Normalize
import pandas as pd
from datetime import datetime, timedelta
import pytz
import numpy as np
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.drawing.image import Image
import os

class HingeNode():
    def __init__(self, time_slice:int, params: DParams,
                 top_temp:float, middle_temp:float, bottom_temp:float, thermocline1:float, thermocline2: float):
        self.time_slice = time_slice
        self.top_temp = top_temp
        self.middle_temp = middle_temp
        self.bottom_temp = bottom_temp
        self.thermocline1 = thermocline1
        self.thermocline2 = thermocline2
        self.params = params
        self.energy = self.get_energy()

    def __repr__(self):
        if self.thermocline2 is not None:
            return f"{self.top_temp}({self.thermocline1}){self.middle_temp}({self.thermocline2}){self.bottom_temp}"
        else:
            return f"{self.top_temp}({self.thermocline1}){self.bottom_temp}"

    def get_energy(self):
        m_layer_kg = self.params.storage_volume*3.785 / self.params.num_layers
        if self.middle_temp is not None:
            kWh_top = (self.thermocline1-0.5)*m_layer_kg * 4.187/3600 * to_kelvin(self.top_temp)
            kWh_midlle = (self.thermocline2-self.thermocline1)*m_layer_kg * 4.187/3600 * to_kelvin(self.middle_temp)
            kWh_bottom = (self.params.num_layers-self.thermocline2+0.5)*m_layer_kg * 4.187/3600 * to_kelvin(self.bottom_temp)
        else:        
            kWh_top = (self.thermocline1-0.5)*m_layer_kg * 4.187/3600 * to_kelvin(self.top_temp)
            kWh_midlle = 0
            kWh_bottom = (self.params.num_layers-self.thermocline1+0.5)*m_layer_kg * 4.187/3600 * to_kelvin(self.bottom_temp)
        return kWh_top + kWh_midlle + kWh_bottom
    
    def plot(self, title=''):  
        norm = Normalize(vmin=self.params.available_top_temps[0]-20, vmax=self.params.available_top_temps[-1]+20)
        cmap = matplotlib.colormaps['Reds'] 
        mt = self.middle_temp if self.middle_temp is not None else self.bottom_temp 
        tank_top_colors = [cmap(norm(x)) for x in [self.top_temp]]
        tank_middle_colors = [cmap(norm(x)) for x in [mt]]
        tank_bottom_colors = [cmap(norm(x)) for x in [self.bottom_temp]]
        th2 = self.thermocline2 if self.thermocline2 is not None else self.thermocline1
        thermocline_reversed = self.params.num_layers - self.thermocline1 + 1
        thermocline2_reversed = self.params.num_layers - th2 + 1
        bars_top = plt.bar([0], [self.thermocline1], bottom=thermocline_reversed, color=tank_top_colors, alpha=0.9, width=0.5)
        bars_middle = plt.bar([0], [th2 - self.thermocline1], bottom=[thermocline2_reversed], color=tank_middle_colors, alpha=0.9, width=0.5)
        bars_bottom = plt.bar([0], [thermocline2_reversed], bottom=[0], color=tank_bottom_colors, alpha=0.9, width=0.5)
        plt.xlim([-1,1])
        plt.xticks([])
        for bar in bars_top:
            plt.text(bar.get_x() + bar.get_width() / 2, bar.get_y() + bar.get_height() / 2, 
                     f'{int(self.top_temp)}', ha='center', va='center', color='white')
        if self.middle_temp is not None:
            for bar in bars_middle:
                plt.text(bar.get_x() + bar.get_width() / 2, bar.get_y() + bar.get_height() / 2, 
                        f'{int(mt)}', ha='center', va='center', color='white')
        for bar in bars_bottom:
            plt.text(bar.get_x() + bar.get_width() / 2, bar.get_y() + bar.get_height() / 2, 
                     f'{int(self.bottom_temp)}', ha='center', va='center', color='white')
        plt.title(title)
        plt.show()


class FloHinge():

    def __init__(self, flo_params: FloParamsHouse0):
        # Graph in which charging is accurate
        self.g = DGraph(flo_params)
        self.g.solve_dijkstra()
        # Graph in which discharging is accurate
        flo_params.DdDeltaTF = 45
        self.dg = DGraph(flo_params)
        self.dg.solve_dijkstra()
        self.hinge_steps: List[HingeNode] = []
        self.start()

    def start(self):
        self.get_hinge_start_state()
        self.evaluate_branches()

    def plot_hinge(self, combo=''):
        # Want to show initial, discharging, and the result of the 3 branches
        sp_time = list(range(self.turn_on_hour+3+2))
        sp_top_temp = []
        sp_middle_temp = []
        sp_bottom_temp = []
        sp_thermocline = []
        sp_thermocline2 = []
        for step in self.hinge_steps:
            sp_top_temp.append(step.top_temp)
            sp_bottom_temp.append(step.bottom_temp)
            sp_thermocline.append(step.thermocline1)
            if step.middle_temp is not None:
                sp_middle_temp.append(step.middle_temp)
                sp_thermocline2.append(step.thermocline2)
            else:
                sp_middle_temp.append(step.bottom_temp)
                sp_thermocline2.append(step.thermocline1)
        # Bottom plot
        norm = Normalize(vmin=self.g.params.available_top_temps[0]-20, vmax=self.g.params.available_top_temps[-1]+20)
        cmap = matplotlib.colormaps['Reds']
        tank_top_colors = [cmap(norm(x)) for x in sp_top_temp]
        tank_middle_colors = [cmap(norm(x)) for x in sp_middle_temp]
        tank_bottom_colors = [cmap(norm(x)) for x in sp_bottom_temp]

        # Reversing thermocline positions
        sp_thermocline_reversed1 = [self.g.params.num_layers - x + 1 for x in sp_thermocline]
        sp_thermocline_reversed2 = [self.g.params.num_layers - x + 1 for x in sp_thermocline2]

        # Stacking the temperatures and thermoclines
        bars_top = plt.bar(sp_time, sp_thermocline, bottom=sp_thermocline_reversed1, color=tank_top_colors, alpha=0.7, width=0.9)
        bars_middle = plt.bar(sp_time, [y-x for x,y in zip(sp_thermocline, sp_thermocline2)], 
                              bottom=sp_thermocline_reversed2, color=tank_middle_colors, alpha=0.7, width=0.9)
        bars_bottom = plt.bar(sp_time, sp_thermocline_reversed2, bottom=0, color=tank_bottom_colors, alpha=0.7, width=0.9)
        plt.xlabel('Time [hours]')
        plt.ylabel('Storage state')
        plt.ylim([0, self.g.params.num_layers])
        plt.yticks([])
        for i, bar in enumerate(bars_top):
            plt.text(bar.get_x() + bar.get_width() / 2, bar.get_y() + bar.get_height() / 2, 
                    f'{int(sp_top_temp[i])}', ha='center', va='center', color='white')
        for i, bar in enumerate(bars_middle):
            if sp_middle_temp[i] != sp_bottom_temp[i]:
                plt.text(bar.get_x() + bar.get_width() / 2, bar.get_y() + bar.get_height() / 2, 
                        f'{int(sp_middle_temp[i])}', ha='center', va='center', color='white')
        for i, bar in enumerate(bars_bottom):
            plt.text(bar.get_x() + bar.get_width() / 2, bar.get_y() + bar.get_height() / 2, 
                    f'{int(sp_bottom_temp[i])}', ha='center', va='center', color='white')
        plt.title(f"{'d-'*self.turn_on_hour}{self.best_combination if combo=='' else combo}-knit")
        plt.savefig('plot_hinge.png', dpi=130)
        # plt.show()

    def get_hinge_start_state(self):
        # Find hour at which the HP is turned on (we trust the decisions to discharge)
        node_i = self.g.initial_node
        self.initial_node = HingeNode(
            time_slice = node_i.time_slice,
            top_temp = node_i.top_temp,
            middle_temp = node_i.middle_temp,
            bottom_temp = node_i.bottom_temp,
            thermocline1 = node_i.thermocline1,
            thermocline2 = node_i.thermocline2,
            params = self.g.params
            )
        self.hinge_steps.append(self.initial_node)
        for i in range(48):
            heat_out = [e.hp_heat_out for e in self.g.edges[node_i] if e.head==node_i.next_node][0]
            if heat_out > 2:
                self.turn_on_hour = i
                break
            node_i = node_i.next_node
        print(f"FLO turns on HP at hour {self.turn_on_hour}")
        # Find what the storage would look like if we discharged until then
        node_i = self.dg.initial_node
        for i in range(self.turn_on_hour):
            node_i = self.dg.edges[node_i][0].head
            self.hinge_steps.append(node_i)
        self.turn_on_node = HingeNode(
            time_slice = node_i.time_slice,
            top_temp = node_i.top_temp,
            middle_temp = node_i.middle_temp,
            bottom_temp = node_i.bottom_temp,
            thermocline1 = node_i.thermocline1,
            thermocline2 = node_i.thermocline2,
            params = self.g.params
            )
        print(f"Estimated storage at the start of hour {self.turn_on_hour}: {self.turn_on_node}")

    def evaluate_branches(self):
        self.feasible_branches = {}
        for branch1_charge in [True, False]:
            for branch2_charge in [True, False]:
                for branch3_charge in [True, False]:
                    combination_name = f"{'C' if branch1_charge else 'D'}-"
                    combination_name += f"{'C' if branch2_charge else 'D'}-"
                    combination_name += f"{'C' if branch3_charge else 'D'}"
                    self.follow_branch(branch1_charge, branch2_charge, branch3_charge, combination_name)
        self.knit_branches()

        for branch in self.feasible_branches:
            print(f"\nCombination: {branch}")
            print(f"- Ends at {self.feasible_branches[branch]['final_state']}")
            print(f"- Knitted to {self.feasible_branches[branch]['knitted_to']}")
            print(f"- Total pathcost: {self.feasible_branches[branch]['total_pathcost']}")

        self.best_combination = min(self.feasible_branches, key=lambda k: self.feasible_branches[k]['total_pathcost'])
        print(f"\nThe best path forward is {self.best_combination}")

        # Find the nodes and pathcosts for hour 1
        if self.turn_on_hour == 0:
            best_combination_starting_with_charge = None
            if [x for x in self.feasible_branches if x.split('-')[0]=='C']:
                best_combination_starting_with_charge = min(
                    [x for x in self.feasible_branches if x.split('-')[0]=='C'], 
                    key=lambda k: self.feasible_branches[k]['total_pathcost']
                    )
            best_combination_starting_with_discharge = None                 
            if [x for x in self.feasible_branches if x.split('-')[0]=='D']:
                best_combination_starting_with_discharge = min(
                    [x for x in self.feasible_branches if x.split('-')[0]=='D'], 
                    key=lambda k: self.feasible_branches[k]['total_pathcost']
                    )
            # Charging pathcost
            cost_of_hour_1 = self.g.params.elec_price_forecast[0] * self.g.params.max_hp_elec_in / 100
            charge1_pathcost = self.feasible_branches[best_combination_starting_with_charge]['total_pathcost'] - cost_of_hour_1
            charge1_node = self.charge_from(self.turn_on_node)
            # Discharging pathcost
            discharge1_pathcost = self.feasible_branches[best_combination_starting_with_discharge]['total_pathcost']
            discharge1_node = self.discharge_from(self.turn_on_node)
        else:
            charge1_pathcost = None
            charge1_node = None
            discharge1_pathcost = self.feasible_branches[self.best_combination]['total_pathcost']
            discharge1_node = self.dg.edges[self.dg.initial_node][0].head
        
        if charge1_node:
            self.charge1_node = DNode(
                time_slice = 0, 
                top_temp=charge1_node.top_temp,
                thermocline1=charge1_node.top_temp,
                parameters=charge1_node.params,
                hinge_node = {
                    'middle_temp': charge1_node.middle_temp,
                    'bottom_temp': charge1_node.bottom_temp,
                    'thermocline2': charge1_node.thermocline2,
                    'pathcost': charge1_pathcost,
                    }
                )
        else:
            self.charge1_node = None
        self.discharge1_node = DNode(
            time_slice = 0, 
            top_temp=discharge1_node.top_temp,
            thermocline1=discharge1_node.top_temp,
            parameters=discharge1_node.params,
            hinge_node = {
                'middle_temp': discharge1_node.middle_temp,
                'bottom_temp': discharge1_node.bottom_temp,
                'thermocline2': discharge1_node.thermocline2,
                'pathcost': discharge1_pathcost,
                }
            )

        for combo in self.feasible_branches:
            if combo != self.best_combination:
                continue
            b1, b2, b3 = [True if x=='C' else False for x in combo.split('-')]
            self.hinge_steps = []
            self.get_hinge_start_state()
            self.follow_branch(b1, b2, b3, combo, final=True)
            self.hinge_steps.append(self.feasible_branches[combo]['knitted_to'])
            self.plot_hinge(combo=combo)


    def follow_branch(self, branch1_charge, branch2_charge, branch3_charge, combination_name, final=False):
        node0 = self.turn_on_node
        total_hinge_cost_usd = 0
        # First hour
        node1 = self.charge_from(node0) if branch1_charge else self.discharge_from(node0)
        if final:
            self.hinge_steps.append(node1)
        h = self.turn_on_node.time_slice
        if branch1_charge:
            total_hinge_cost_usd += self.g.params.elec_price_forecast[h] * self.g.params.max_hp_elec_in / 100
        else:
            RSWT = self.g.params.rswt_forecast[h]
            if node0.top_temp < RSWT or node1.top_temp < RSWT - self.g.params.delta_T(RSWT):
                return
        # Second hour
        node2 = self.charge_from(node1) if branch2_charge else self.discharge_from(node1)
        if final:
            self.hinge_steps.append(node2)
        h += 1
        if branch2_charge:
            total_hinge_cost_usd += self.g.params.elec_price_forecast[h] * self.g.params.max_hp_elec_in / 100
        else:
            RSWT = self.g.params.rswt_forecast[h]
            if node1.top_temp < RSWT or node2.top_temp < RSWT - self.g.params.delta_T(RSWT):
                return
        # Third hour
        node3 = self.charge_from(node2) if branch3_charge else self.discharge_from(node2)
        if final:
            self.hinge_steps.append(node3)
        h += 1
        if branch3_charge:
            total_hinge_cost_usd += self.g.params.elec_price_forecast[h] * self.g.params.max_hp_elec_in / 100
        else:
            RSWT = self.g.params.rswt_forecast[h]
            if node2.top_temp < RSWT or node3.top_temp < RSWT - self.g.params.delta_T(RSWT):
                return
        # Add to feasible branches
        if not final:
            self.feasible_branches[combination_name] = {'hinge_cost': total_hinge_cost_usd, 'final_state': node3}


    def discharge_from(self, n: HingeNode):
        next_node_top_temp = n.top_temp
        next_node_energy = n.energy - self.g.params.load_forecast[n.time_slice]
        
        if n.top_temp - self.dg.params.delta_T(n.top_temp) < n.bottom_temp or n.middle_temp is not None:
            flo_params_temporary: FloParamsHouse0 = self.dg.params.config.model_copy()
            flo_params_temporary.HorizonHours = 24
            flo_params_temporary.InitialTopTempF = n.top_temp if n.top_temp<=175 else 175
            flo_params_temporary.InitialBottomTempF = n.bottom_temp if n.middle_temp is None else n.middle_temp
            flo_params_temporary.InitialThermocline = n.thermocline1 if n.thermocline2 is None else (self.dg.params.num_layers-n.thermocline2+n.thermocline1)
            temporary_g = DGraph(flo_params_temporary)
            node_after = min(temporary_g.nodes[0], key=lambda x: abs(x.energy-next_node_energy))
            next_node_top_temp = node_after.top_temp
            next_node_middle_temp = node_after.middle_temp
            next_node_bottom_temp = node_after.bottom_temp
            next_node_thermocline = node_after.thermocline1
            next_node_thermocline2 = node_after.thermocline2
        else:
            temporary_g = None
            next_node_middle_temp = None
            next_node_bottom_temp = n.bottom_temp
            next_node_thermocline2 = None
            # Find thermocline position such that kWh_top + kWh_bottom = next_node_energy
            m_layer_kg = self.g.params.storage_volume*3.785 / self.g.params.num_layers       
            top, bottom = to_kelvin(next_node_top_temp), to_kelvin(next_node_bottom_temp)
            A = m_layer_kg * 4.187/3600
            next_node_thermocline = int(1/(top-bottom) * (next_node_energy/A - (-0.5*top + (self.g.params.num_layers+0.5)*bottom)))
            while next_node_thermocline < 1:
                next_node_top_temp = next_node_bottom_temp
                next_node_bottom_temp = round(next_node_bottom_temp - self.g.params.delta_T(next_node_bottom_temp))
                top, bottom = to_kelvin(next_node_top_temp), to_kelvin(next_node_bottom_temp)
                next_node_thermocline = int(1/(top-bottom) * (next_node_energy/A - (-0.5*top + (self.g.params.num_layers+0.5)*bottom)))

        next_node = HingeNode(
            time_slice = n.time_slice+1,
            top_temp = next_node_top_temp,
            middle_temp = next_node_middle_temp,
            bottom_temp = next_node_bottom_temp,
            thermocline1 = next_node_thermocline,
            thermocline2 = next_node_thermocline2,
            params = self.g.params
        )
        return next_node

    def charge_from(self, n: HingeNode):
        next_node_bottom_temp = n.bottom_temp
        load = self.g.params.load_forecast[n.time_slice]
        hp = self.g.params.max_hp_elec_in * self.g.params.COP(self.g.params.oat_forecast[n.time_slice], 0)
        heat_to_store = hp - load
        next_node_energy = n.energy + heat_to_store

        if n.bottom_temp + self.g.params.delta_T(n.bottom_temp) < n.top_temp:
            if n.middle_temp is not None:
                top_mixed = (n.top_temp*n.thermocline1 + n.middle_temp*(n.thermocline2-n.thermocline1))/n.thermocline2
                next_node_top_temp = round(
                    (top_mixed*n.thermocline2 + (n.bottom_temp+self.g.params.delta_T(n.bottom_temp))*(self.g.params.num_layers-n.thermocline2))/self.g.params.num_layers
                    )
            else:
                next_node_top_temp = round(
                    n.thermocline1/self.g.params.num_layers * n.top_temp 
                    + (self.g.params.num_layers-n.thermocline1)/self.g.params.num_layers * (n.bottom_temp + self.g.params.delta_T(n.bottom_temp))
                    )
        else:
            next_node_top_temp = n.top_temp

        # Find thermocline position such that kWh_top + kWh_bottom = next_node_energy
        m_layer_kg = self.g.params.storage_volume*3.785 / self.g.params.num_layers       
        top, bottom = to_kelvin(next_node_top_temp), to_kelvin(next_node_bottom_temp)
        A = m_layer_kg * 4.187/3600
        next_node_thermocline = int(1/(top-bottom) * (next_node_energy/A - (-0.5*top + (self.g.params.num_layers+0.5)*bottom)))

        while next_node_thermocline > self.g.params.num_layers:
            next_node_bottom_temp = next_node_top_temp
            next_node_top_temp = round(next_node_top_temp + self.g.params.delta_T(next_node_top_temp))
            top, bottom = to_kelvin(next_node_top_temp), to_kelvin(next_node_bottom_temp)
            next_node_thermocline = int(1/(top-bottom) * (next_node_energy/A - (-0.5*top + (self.g.params.num_layers+0.5)*bottom)))

        next_node = HingeNode(
            time_slice = n.time_slice+1,
            top_temp = next_node_top_temp,
            middle_temp = None,
            bottom_temp = next_node_bottom_temp,
            thermocline1 = next_node_thermocline,
            thermocline2 = None,
            params = self.g.params
        )
        return next_node

    def knit_branches(self):
        for branch in self.feasible_branches:
            n: HingeNode = self.feasible_branches[branch]['final_state']
            knitted_node = [min(self.g.nodes[n.time_slice], key= lambda x: abs(x.energy-n.energy))][0]            
            self.feasible_branches[branch]['knitted_to'] = knitted_node
            self.feasible_branches[branch]['total_pathcost'] = round(knitted_node.pathcost + self.feasible_branches[branch]['hinge_cost'],2)

    def generate_bid(self):
        # add new nodes and edges
        if self.charge1_node:
            self.g.nodes[0].extend([self.charge1_node, self.discharge1_node])
            # Charge edge
            charge1_cost = self.g.params.elec_price_forecast[0] * self.g.params.max_hp_elec_in / 100
            charge1_hp_heat_out = self.g.params.max_hp_elec_in * self.g.params.COP(self.g.params.oat_forecast[0], 0) 
            charge1_edge = DEdge(self.g.initial_node, self.charge1_node, charge1_cost, charge1_hp_heat_out)
            self.g.edges[self.g.initial_node].append(charge1_edge)
            # Discharge edge
            discharge1_edge = DEdge(self.g.initial_node, self.discharge1_node, 0, 0)
            self.g.edges[self.g.initial_node].append(discharge1_edge)
        else:
            self.g.nodes[0].extend([self.discharge1_node])
            # Discharge edge
            discharge1_edge = DEdge(self.g.initial_node, self.discharge1_node, 0, 0)
            self.g.edges[self.g.initial_node].append(discharge1_edge)

        self.pq_pairs: List[PriceQuantityUnitless] = []
        forecasted_price_usd_mwh = self.g.params.elec_price_forecast[0] * 10
        # For every possible price
        min_elec_ctskwh, max_elec_ctskwh = -10, 200
        for elec_price_usd_mwh in sorted(list(range(min_elec_ctskwh*10, max_elec_ctskwh*10))+[forecasted_price_usd_mwh]):
            # Update the fake cost of initial node edges with the selected price
            for edge in self.g.edges[self.g.initial_node]:
                if edge.cost >= 1e4: # penalized node
                    edge.fake_cost = edge.cost
                elif edge.rswt_minus_edge_elec is not None: # penalized node
                    edge.fake_cost = edge.rswt_minus_edge_elec * elec_price_usd_mwh/1000
                else:
                    cop = self.g.params.COP(oat=self.g.params.oat_forecast[0], lwt=edge.head.top_temp)
                    edge.fake_cost = edge.hp_heat_out / cop * elec_price_usd_mwh/1000
            # Find the best edge with the given price
            best_edge: DEdge = min(self.g.edges[self.g.initial_node], key=lambda e: e.head.pathcost + e.fake_cost)
            if best_edge.hp_heat_out < 0: 
                best_edge_neg = max([e for e in self.g.edges[self.g.initial_node] if e.hp_heat_out<0], key=lambda e: e.hp_heat_out)
                best_edge_pos = min([e for e in self.g.edges[self.g.initial_node] if e.hp_heat_out>=0], key=lambda e: e.hp_heat_out)
                best_edge = best_edge_pos if (-best_edge_neg.hp_heat_out >= best_edge_pos.hp_heat_out) else best_edge_neg
            # Find the associated quantity
            cop = self.g.params.COP(oat=self.g.params.oat_forecast[0], lwt=best_edge.head.top_temp)
            best_quantity_kwh = best_edge.hp_heat_out / cop
            best_quantity_kwh = 0 if best_quantity_kwh<0 else best_quantity_kwh
            if not self.pq_pairs:
                self.pq_pairs.append(
                    PriceQuantityUnitless(
                        PriceTimes1000 = int(elec_price_usd_mwh * 1000),
                        QuantityTimes1000 = int(best_quantity_kwh * 1000))
                )
            else:
                # Record a new pair if at least 0.01 kWh of difference in quantity with the previous one
                if self.pq_pairs[-1].QuantityTimes1000 - int(best_quantity_kwh * 1000) > 10:
                    self.pq_pairs.append(
                        PriceQuantityUnitless(
                            PriceTimes1000 = int(elec_price_usd_mwh * 1000),
                            QuantityTimes1000 = int(best_quantity_kwh * 1000))
                    )
        # remove new nodes and edges
        if self.charge1_node:
            self.g.nodes[0].remove(self.charge1_node)
            self.g.edges[self.g.initial_node].remove(charge1_edge)
        self.g.nodes[0].remove(self.discharge1_node)
        self.g.edges[self.g.initial_node].remove(discharge1_edge)

        return self.pq_pairs
    
    def export_to_excel(self):
        # Sort nodes by energy and assign an index
        for time_slice in range(self.g.params.horizon+1):
            self.g.nodes_by_energy = sorted(self.g.nodes[time_slice], key=lambda x: (x.energy, x.top_temp), reverse=True)
            for n in self.g.nodes[time_slice]:
                n.index = self.g.nodes_by_energy.index(n)+1

        # Along the shortest path
        electricitiy_used, heat_delivered = [], []
        node_i = self.g.initial_node
        while node_i.next_node is not None:
            losses = self.g.params.storage_losses_percent/100 * (node_i.energy-self.g.bottom_node.energy)
            if self.g.params.load_forecast[node_i.time_slice]==0 and losses>0 and losses<self.g.params.energy_between_nodes[node_i.top_temp]:
                losses = self.g.params.energy_between_nodes[node_i.top_temp] + 1/1e9
            store_heat_in = node_i.next_node.energy - node_i.energy
            hp_heat_out = store_heat_in + self.g.params.load_forecast[node_i.time_slice] + losses
            cop = self.g.params.COP(oat=self.g.params.oat_forecast[node_i.time_slice], lwt=node_i.next_node.top_temp)
            heat_delivered.append(hp_heat_out)
            electricitiy_used.append(hp_heat_out/cop)
            node_i = node_i.next_node
        
        # First dataframe: the Dijkstra graph
        dijkstra_pathcosts = {}
        dijkstra_pathcosts['Model'] = [repr(x) for x in self.g.nodes_by_energy]
        dijkstra_pathcosts['Energy (relative)'] = [round(x.energy-self.g.bottom_node.energy,2) for x in self.g.nodes_by_energy]
        dijkstra_pathcosts['Index'] = list(range(1,len(self.g.nodes_by_energy)+1))
        dijkstra_nextnodes = dijkstra_pathcosts.copy()
        for h in range(self.g.params.horizon):
            dijkstra_pathcosts[h] = [round(x.pathcost,2) for x in sorted(self.g.nodes[h], key=lambda x: x.index)]
            dijkstra_nextnodes[h] = [x.next_node.index for x in sorted(self.g.nodes[h], key=lambda x: x.index)]
        dijkstra_pathcosts[self.g.params.horizon] = [0 for x in self.g.nodes[self.g.params.horizon]]
        dijkstra_nextnodes[self.g.params.horizon] = [np.nan for x in self.g.nodes[self.g.params.horizon]]
        dijkstra_pathcosts_df = pd.DataFrame(dijkstra_pathcosts)
        dijkstra_nextnodes_df = pd.DataFrame(dijkstra_nextnodes)
        
        # Second dataframe: the forecasts
        start_time = datetime.fromtimestamp(self.g.params.start_time, tz=pytz.timezone("America/New_York"))
        forecast_df = pd.DataFrame({'Forecast':['0'], 'Unit':['0'], **{h: [0.0] for h in range(self.g.params.horizon)}})
        forecast_df.loc[0] = ['Hour'] + [start_time.strftime("%d/%m/%Y")] + [(start_time + timedelta(hours=x)).hour for x in range(self.g.params.horizon)]
        forecast_df.loc[1] = ['Price - total'] + ['cts/kWh'] + self.g.params.elec_price_forecast
        forecast_df.loc[2] = ['Price - distribution'] + ['cts/kWh'] + self.g.params.dist_forecast
        forecast_df.loc[3] = ['Price - LMP'] + ['cts/kWh'] + self.g.params.lmp_forecast
        forecast_df.loc[4] = ['Heating load'] + ['kW'] + [round(x,2) for x in self.g.params.load_forecast]
        forecast_df.loc[5] = ['OAT'] + ['F'] + [round(x,2) for x in self.g.params.oat_forecast]
        forecast_df.loc[6] = ['Required SWT'] + ['F'] + [round(x) for x in self.g.params.rswt_forecast]
        
        # Third dataframe: the shortest path
        shortestpath_df = pd.DataFrame({'Shortest path':['0'], 'Unit':['0'], **{h: [0.0] for h in range(self.g.params.horizon+1)}})
        shortestpath_df.loc[0] = ['Electricity used'] + ['kWh'] + [round(x,3) for x in electricitiy_used] + [0]
        shortestpath_df.loc[1] = ['Heat delivered'] + ['kWh'] + [round(x,3) for x in heat_delivered] + [0]
        shortestpath_df.loc[2] = ['Cost - total'] + ['cts'] + [round(x*y,2) for x,y in zip(electricitiy_used, self.g.params.elec_price_forecast)] + [0]
        shortestpath_df.loc[3] = ['Cost - distribution'] + ['cts'] + [round(x*y,2) for x,y in zip(electricitiy_used, self.g.params.dist_forecast)] + [0]
        shortestpath_df.loc[4] = ['Cost - LMP'] + ['cts'] + [round(x*y,2) for x,y in zip(electricitiy_used, self.g.params.lmp_forecast)] + [0]
        
        # Fourth dataframe: the results
        total_usd = round(self.g.initial_node.pathcost,2)
        total_elec = round(sum(electricitiy_used),2)
        total_heat = round(sum(heat_delivered),2)
        next_index = self.g.initial_node.next_node.index
        results = ['Cost ($)', total_usd, 'Electricity (kWh)', total_elec, 'Heat (kWh)', total_heat, 'Next step index', next_index]
        results_df = pd.DataFrame({'RESULTS':results})
        
        # Highlight shortest path
        highlight_positions = []
        node_i = self.g.initial_node
        while node_i.next_node is not None:
            highlight_positions.append((node_i.index+len(forecast_df)+len(shortestpath_df)+2, 3+node_i.time_slice))
            node_i = node_i.next_node
        highlight_positions.append((node_i.index+len(forecast_df)+len(shortestpath_df)+2, 3+node_i.time_slice))
        
        # Add the parameters to a seperate sheet
        parameters = self.g.params.config.to_dict()
        parameters_df = pd.DataFrame(list(parameters.items()), columns=['Variable', 'Value'])

        # Add the PQ pairs to a seperate sheet and plot the curve
        pq_pairs = self.generate_bid()
        print(pq_pairs)
        print(len(pq_pairs))
        prices = [x.PriceTimes1000 for x in pq_pairs]
        quantities = [x.QuantityTimes1000/1000 for x in pq_pairs]
        pqpairs_df = pd.DataFrame({'price':[x/1000 for x in prices], 'quantity':quantities})
        # To plot quantities on x-axis and prices on y-axis
        ps, qs = [], []
        index_p = 0
        expected_price_usd_mwh = self.g.params.elec_price_forecast[0] * 10
        for p in sorted(list(range(min(prices), max(prices)+1)) + [expected_price_usd_mwh*1000]):
            ps.append(p/1000)
            if index_p+1 < len(prices) and p >= prices[index_p+1]:
                index_p += 1
            if p == expected_price_usd_mwh*1000:
                interesection = (quantities[index_p], expected_price_usd_mwh)
            qs.append(quantities[index_p])
        plt.plot(qs, ps, label='demand (bid)')
        prices = [x.PriceTimes1000/1000 for x in pq_pairs]
        plt.scatter(quantities, prices)
        plt.plot([min(quantities)-1, max(quantities)+1],[expected_price_usd_mwh]*2, label="supply (expected market price)")
        plt.scatter(interesection[0], interesection[1])
        plt.text(interesection[0]+0.25, interesection[1]+15, f'({round(interesection[0],3)}, {round(interesection[1],1)})', fontsize=10, color='tab:orange')
        plt.xticks(quantities)
        if min([abs(x-expected_price_usd_mwh) for x in prices]) < 5:
            plt.yticks(prices)
        else:
            plt.yticks(prices + [expected_price_usd_mwh])
        plt.ylabel("Price [USD/MWh]")
        plt.xlabel("Quantity [kWh]")
        plt.grid(alpha=0.3)
        plt.legend()
        plt.tight_layout()
        plt.savefig('plot_pq.png', dpi=130)
        plt.close()

        # Write to Excel
        os.makedirs('results', exist_ok=True)
        start = datetime.fromtimestamp(self.g.params.start_time, tz=pytz.timezone("America/New_York")).strftime('%Y-%m-%d %H:%M')
        # file_path = os.path.join('results', f'result_{start}.xlsx')
        file_path = 'result.xlsx'
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            results_df.to_excel(writer, index=False, sheet_name='Pathcost')
            results_df.to_excel(writer, index=False, sheet_name='Next node')
            forecast_df.to_excel(writer, index=False, startcol=1, sheet_name='Pathcost')
            forecast_df.to_excel(writer, index=False, startcol=1, sheet_name='Next node')
            shortestpath_df.to_excel(writer, index=False, startcol=1, startrow=len(forecast_df)+1, sheet_name='Pathcost')
            shortestpath_df.to_excel(writer, index=False, startcol=1, startrow=len(forecast_df)+1, sheet_name='Next node')
            dijkstra_pathcosts_df.to_excel(writer, index=False, startrow=len(forecast_df)+len(shortestpath_df)+2, sheet_name='Pathcost')
            dijkstra_nextnodes_df.to_excel(writer, index=False, startrow=len(forecast_df)+len(shortestpath_df)+2, sheet_name='Next node')
            parameters_df.to_excel(writer, index=False, sheet_name='Parameters')
            
            # Add plot in a seperate sheet
            self.g.plot(show=False)
            plot_sheet = writer.book.create_sheet(title='Plot')
            plot_sheet.add_image(Image('plot.png'), 'A1')
            plot_sheet.add_image(Image('plot_hinge.png'), 'T1')

            # Add plot in a seperate sheet
            plot2_sheet = writer.book.create_sheet(title='PQ pairs')
            pqpairs_df.to_excel(writer, index=False, sheet_name='PQ pairs')
            plot2_sheet.add_image(Image('plot_pq.png'), 'C1')

            # Layout
            pathcost_sheet = writer.sheets['Pathcost']
            nextnode_sheet = writer.sheets['Next node']
            parameters_sheet = writer.sheets['Parameters']
            for row in pathcost_sheet['A1:A10']:
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
            for row in nextnode_sheet['A1:A10']:
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
            for row in parameters_sheet[f'B1:B{len(parameters_df)+1}']:
                for cell in row:
                    cell.alignment = Alignment(horizontal='right')
            pathcost_sheet.column_dimensions['A'].width = 17.5
            pathcost_sheet.column_dimensions['B'].width = 15
            pathcost_sheet.column_dimensions['C'].width = 15
            nextnode_sheet.column_dimensions['A'].width = 17.5
            nextnode_sheet.column_dimensions['B'].width = 15
            nextnode_sheet.column_dimensions['C'].width = 15
            parameters_sheet.column_dimensions['A'].width = 40
            parameters_sheet.column_dimensions['B'].width = 70
            pathcost_sheet.freeze_panes = 'D16'
            nextnode_sheet.freeze_panes = 'D16'

            # Highlight shortest path
            highlight_fill = PatternFill(start_color='72ba93', end_color='72ba93', fill_type='solid')
            for row in range(len(forecast_df)+len(shortestpath_df)+2):
                pathcost_sheet.cell(row=row+1, column=1).fill = highlight_fill
                nextnode_sheet.cell(row=row+1, column=1).fill = highlight_fill
            for row, col in highlight_positions:
                pathcost_sheet.cell(row=row+1, column=col+1).fill = highlight_fill
                nextnode_sheet.cell(row=row+1, column=col+1).fill = highlight_fill

        os.remove('plot.png')        
        os.remove('plot_pq.png')
        os.remove('plot_hinge.png')


if __name__ == '__main__':

    # ----------------------------------
    # Load a FLO params
    def from_dict_msg(data):
        message = MessageSql(
                message_id=data["MessageId"],
                from_alias=data["FromAlias"],
                message_type_name=data["MessageTypeName"],
                message_persisted_ms=data["MessagePersistedMs"],
                payload=data["Payload"],
                message_created_ms=data.get("MessageCreatedMs")  # This is optional
            )
        return message
    with open('messages.json', 'r') as file:
        messages_dict = json.load(file)
    message_loaded = from_dict_msg(messages_dict)
    flo_params = FloParamsHouse0(**message_loaded.payload)
    # ----------------------------------
    
    f = FloHinge(flo_params)

    print('\n')
    print(f"Initial state: {f.initial_node}")
    print(f"Charging gives: {f.charge1_node}")
    print(f"Discharging gives: {f.discharge1_node}")
    print('\n')

    f.generate_bid()
    f.export_to_excel()