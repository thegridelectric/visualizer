from flo import DGraph, DParams, DNode, DEdge, to_kelvin
from named_types import FloParamsHouse0, PriceQuantityUnitless
from models import MessageSql
import json
from typing import List, Union
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.colors import Normalize
import pandas as pd
from datetime import datetime, timedelta
import pytz
import numpy as np
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.drawing.image import Image
import os

PRINT = False


class HingeNode():
    def __init__(self, time_slice:int, params: DParams, top_temp:float, middle_temp:float, bottom_temp:float, 
                 thermocline1:float, thermocline2: float, pathcost: float=None):
        self.time_slice = time_slice
        self.top_temp = top_temp
        self.middle_temp = middle_temp
        self.bottom_temp = bottom_temp
        self.thermocline1 = thermocline1
        self.thermocline2 = thermocline2
        self.params = params
        self.pathcost = pathcost
        self.energy = self.get_energy()

    def __repr__(self):
        if self.thermocline2 is not None:
            return f"{self.top_temp}({self.thermocline1}){self.middle_temp}({self.thermocline2}){self.bottom_temp}"
        else:
            return f"{self.top_temp}({self.thermocline1}){self.bottom_temp}"

    def get_energy(self):
        m_layer_kg = self.params.storage_volume*3.785 / self.params.num_layers
        if self.middle_temp is not None:
            kWh_top = (self.thermocline1-0.5)*m_layer_kg * 4.187/3600 * to_kelvin(self.top_temp)
            kWh_midlle = (self.thermocline2-self.thermocline1)*m_layer_kg * 4.187/3600 * to_kelvin(self.middle_temp)
            kWh_bottom = (self.params.num_layers-self.thermocline2+0.5)*m_layer_kg * 4.187/3600 * to_kelvin(self.bottom_temp)
        else:        
            kWh_top = (self.thermocline1-0.5)*m_layer_kg * 4.187/3600 * to_kelvin(self.top_temp)
            kWh_midlle = 0
            kWh_bottom = (self.params.num_layers-self.thermocline1+0.5)*m_layer_kg * 4.187/3600 * to_kelvin(self.bottom_temp)
        return kWh_top + kWh_midlle + kWh_bottom
    
    def plot(self, title=''):  
        norm = Normalize(vmin=self.params.available_top_temps[0]-20, vmax=self.params.available_top_temps[-1]+20)
        cmap = matplotlib.colormaps['Reds'] 
        mt = self.middle_temp if self.middle_temp is not None else self.bottom_temp 
        tank_top_colors = [cmap(norm(x)) for x in [self.top_temp]]
        tank_middle_colors = [cmap(norm(x)) for x in [mt]]
        tank_bottom_colors = [cmap(norm(x)) for x in [self.bottom_temp]]
        th2 = self.thermocline2 if self.thermocline2 is not None else self.thermocline1
        thermocline_reversed = self.params.num_layers - self.thermocline1 + 1
        thermocline2_reversed = self.params.num_layers - th2 + 1
        bars_top = plt.bar([0], [self.thermocline1], bottom=thermocline_reversed, color=tank_top_colors, alpha=0.9, width=0.5)
        bars_middle = plt.bar([0], [th2 - self.thermocline1], bottom=[thermocline2_reversed], color=tank_middle_colors, alpha=0.9, width=0.5)
        bars_bottom = plt.bar([0], [thermocline2_reversed], bottom=[0], color=tank_bottom_colors, alpha=0.9, width=0.5)
        plt.xlim([-1,1])
        plt.xticks([])
        for bar in bars_top:
            plt.text(bar.get_x() + bar.get_width() / 2, bar.get_y() + bar.get_height() / 2, 
                     f'{int(self.top_temp)}', ha='center', va='center', color='white')
        if self.middle_temp is not None:
            for bar in bars_middle:
                plt.text(bar.get_x() + bar.get_width() / 2, bar.get_y() + bar.get_height() / 2, 
                        f'{int(mt)}', ha='center', va='center', color='white')
        for bar in bars_bottom:
            plt.text(bar.get_x() + bar.get_width() / 2, bar.get_y() + bar.get_height() / 2, 
                     f'{int(self.bottom_temp)}', ha='center', va='center', color='white')
        plt.title(title)
        plt.show()


class FloHinge():

    def __init__(self, flo_params: FloParamsHouse0, hinge_hours: int, num_nodes: List[int]):
        self.flo_params = flo_params
        self.hinge_hours = hinge_hours
        self.num_nodes = num_nodes
        self.g = DGraph(flo_params)
        self.g.solve_dijkstra()
        flo_params.DdDeltaTF = flo_params.DischargingDdDeltaTF
        self.dg = DGraph(flo_params)
        self.num_layers = self.g.params.num_layers
        self.start()

    def start(self):
        self.initial_node = self.to_hingenode(self.dg.initial_node)
        self.hinge_steps: List[HingeNode] = [self.initial_node]
        print(f"Estimated storage at the start: {self.initial_node}")

        # Find the HP max thermal output for every hour in the hinge
        hp_max_kwh_th = [
            round(self.flo_params.HpMaxElecKw * self.g.params.COP(self.g.params.oat_forecast[x]),2)
            for x in range(self.hinge_hours)
            ]
        # Create a list of all available HP thermal outputs for every hour in the hinge
        step_size_kwh = [hp_max_kwh_th[x] / (self.num_nodes[x]-1) for x in range(self.hinge_hours)]
        available_paths_kwh = [[round(i * step_size_kwh[x],2) for i in range(self.num_nodes[x])] for x in range(self.hinge_hours)]   
        # Remove the non-zero element that is closest to the load and add the load
        for i in range(self.hinge_hours):
            if self.num_nodes[i] > 2:
                load = round(self.g.params.load_forecast[i],2)
                closest_to_load = min([x for x in available_paths_kwh[i] if x>0], key=lambda x: abs(x-load))
                available_paths_kwh[i].remove(closest_to_load)
                available_paths_kwh[i] = sorted(available_paths_kwh[i] + [load])
        self.available_paths_kwh = available_paths_kwh

        # Check the number of available possibilities
        num_combinations = 1
        for h in range(self.hinge_hours):
            num_combinations *= len(available_paths_kwh[h])
            print(f"Hour {h} options: {available_paths_kwh[h]} kWh_th")
        print(f"There are {num_combinations} possible combinations")

        # Explore all possibilities
        self.feasible_branches = {}
        from itertools import product
        for combination in product(*available_paths_kwh):
            self.create_branch(list(combination))
        self.knit_branches()

        if PRINT:
            for branch in self.feasible_branches:
                print(f"\nCombination: {branch}")
                print(f"- Ends at {self.feasible_branches[branch]['final_state']}")
                print(f"- Knitted to {self.feasible_branches[branch]['knitted_to']}")
                print(f"- Total pathcost: {self.feasible_branches[branch]['total_pathcost']}")

        # Best combination
        self.best_combination = min(self.feasible_branches, key=lambda k: self.feasible_branches[k]['total_pathcost'])
        print(f"\nThe best path forward is {self.best_combination}")
        self.hinge_steps = [self.initial_node]
        self.create_branch(self.best_combination, best_combination=True)
        self.hinge_steps.append(self.feasible_branches[self.best_combination]['knitted_to'])
        self.plot_hinge()
        self.quick_plot()

    def create_branch(self, combination, best_combination=False):
        node = self.initial_node
        load = [round(x,2) for x in self.g.params.load_forecast]
        cop = [self.g.params.COP(self.g.params.oat_forecast[x]) for x in range(self.hinge_hours)]
        elec_price = [x/100 for x in self.g.params.elec_price_forecast]
        branch_cost = 0
        for h in range(self.hinge_hours):
            branch_cost += combination[h] / cop[h] * elec_price[h]
            heat_to_store = combination[h]-load[h]
            if heat_to_store > 0:
                node = self.charge(node, heat_to_store)
                if node.top_temp > 175:
                    return
            elif heat_to_store < 0:
                node_before = node
                node = self.discharge(node, -heat_to_store)
                rswt = self.g.params.rswt_forecast[h]
                if node_before.top_temp < rswt or node.top_temp < rswt - self.g.params.delta_T(rswt):
                    return
            else:
                node = self.to_hingenode(node, time_slice=node.time_slice+1)
            if best_combination:
                self.hinge_steps.append(node)
        if not best_combination:
            self.feasible_branches[tuple(combination)] = {
                'branch_cost': round(branch_cost,3), 
                'final_state': node
                }

    def discharge(self, n: HingeNode, discharge_kwh: float):
        next_node_energy = n.energy - discharge_kwh
        if n.top_temp - self.dg.params.delta_T(n.top_temp) < n.bottom_temp or n.middle_temp is not None:
            # Build a new discharging graph from current node and find the node that matches the next node energy
            flo_params_temporary: FloParamsHouse0 = self.dg.params.config.model_copy()
            flo_params_temporary.HorizonHours = 1
            flo_params_temporary.InitialTopTempF = n.top_temp if n.top_temp<=175 else 175
            flo_params_temporary.InitialBottomTempF = n.bottom_temp if n.middle_temp is None else n.middle_temp
            flo_params_temporary.InitialBottomTempF = flo_params_temporary.InitialBottomTempF if flo_params_temporary.InitialBottomTempF<=170 else 170
            flo_params_temporary.InitialThermocline = n.thermocline1 if n.thermocline2 is None else (self.dg.params.num_layers-n.thermocline2+n.thermocline1)
            temporary_g = DGraph(flo_params_temporary)
            node_after = min(temporary_g.nodes[0], key=lambda x: abs(x.energy-next_node_energy))
            return self.to_hingenode(node_after, time_slice=n.time_slice+1)
        else:
            # Starting with current top and bottom, find the thermocline position that matches the next node energy
            next_node_top_temp = n.top_temp
            next_node_bottom_temp = n.bottom_temp
            next_node_thermocline = self.find_thermocline(next_node_top_temp, next_node_bottom_temp, next_node_energy)
            while next_node_thermocline < 1:
                next_node_top_temp = next_node_bottom_temp
                next_node_bottom_temp = round(next_node_bottom_temp - self.g.params.delta_T(next_node_bottom_temp))
                next_node_thermocline = self.find_thermocline(next_node_top_temp, next_node_bottom_temp, next_node_energy)
            return HingeNode(
                time_slice = n.time_slice+1,
                top_temp = next_node_top_temp,
                middle_temp = None,
                bottom_temp = next_node_bottom_temp,
                thermocline1 = next_node_thermocline,
                thermocline2 = None,
                params = self.g.params
            )
    
    def charge(self, n: HingeNode, charge_kwh: float):
        next_node_energy = n.energy + charge_kwh
        if n.bottom_temp + self.g.params.delta_T(n.bottom_temp) < n.top_temp:
            # The next top temperature will be a mix of the current top and bottom+deltaT
            if n.middle_temp is not None:
                top_mixed = (n.top_temp*n.thermocline1 + n.middle_temp*(n.thermocline2-n.thermocline1))/n.thermocline2
                next_node_top_temp = round(
                    (top_mixed*n.thermocline2 
                     + (n.bottom_temp+self.g.params.delta_T(n.bottom_temp))*(self.num_layers-n.thermocline2))/self.num_layers
                    )
            else:
                next_node_top_temp = round(
                    (n.top_temp*n.thermocline1
                     + (n.bottom_temp+self.g.params.delta_T(n.bottom_temp))*(self.num_layers-n.thermocline1))/self.num_layers
                    )
        else:
            next_node_top_temp = n.top_temp
        # Starting with that top and current bottom, find the thermocline position that matches the next node energy
        next_node_bottom_temp = n.bottom_temp
        next_node_thermocline = self.find_thermocline(next_node_top_temp, next_node_bottom_temp, next_node_energy)
        while next_node_thermocline > self.num_layers:
            next_node_bottom_temp = next_node_top_temp
            next_node_top_temp = round(next_node_top_temp + self.g.params.delta_T(next_node_top_temp))
            next_node_thermocline = self.find_thermocline(next_node_top_temp, next_node_bottom_temp, next_node_energy)
        return HingeNode(
            time_slice = n.time_slice+1,
            top_temp = next_node_top_temp,
            middle_temp = None,
            bottom_temp = next_node_bottom_temp,
            thermocline1 = next_node_thermocline,
            thermocline2 = None,
            params = self.g.params
        )
        
    def find_thermocline(self, top_temp, bottom_temp, energy):
        top, bottom = to_kelvin(top_temp), to_kelvin(bottom_temp)
        m_layer_kg = self.g.params.storage_volume*3.785 / self.g.params.num_layers      
        return int(1/(top-bottom)*(energy/(m_layer_kg*4.187/3600)-(-0.5*top+(self.num_layers+0.5)*bottom)))

    def knit_branches(self):
        for branch in self.feasible_branches:
            n: HingeNode = self.feasible_branches[branch]['final_state']
            knitted_node = [min(self.g.nodes[n.time_slice], key= lambda x: abs(x.energy-n.energy))][0]            
            self.feasible_branches[branch]['knitted_to'] = knitted_node
            self.feasible_branches[branch]['total_pathcost'] = round(knitted_node.pathcost + self.feasible_branches[branch]['branch_cost'],2)

    def generate_bid(self):
        # Add new nodes and edges
        hinge_hour0_edges: List[DEdge] = []
        for hour0_kwh in self.available_paths_kwh[0]:
            load0_kwh = self.g.params.load_forecast[0]
            heat_to_store = hour0_kwh-load0_kwh
            if heat_to_store > 0:
                node = self.charge(self.initial_node, heat_to_store)
                if node.top_temp > 175:
                    continue
            elif heat_to_store < 0:
                node = self.discharge(self.initial_node, -heat_to_store)
                rswt = self.g.params.rswt_forecast[0]
                if self.initial_node.top_temp < rswt or node.top_temp < rswt - self.g.params.delta_T(rswt):
                    continue
            else:
                node = self.to_hingenode(self.initial_node, time_slice=1)
            hour0_cost = hour0_kwh / self.g.params.COP(self.g.params.oat_forecast[0]) * self.g.params.elec_price_forecast[0]/100
            if [x for x in self.feasible_branches if x[0]==hour0_kwh]:
                best_branch_from_hour1 = min(
                    [x for x in self.feasible_branches if x[0]==hour0_kwh],
                    key=lambda k: self.feasible_branches[k]['total_pathcost']
                    )
                pathcost = self.feasible_branches[best_branch_from_hour1]['total_pathcost'] - hour0_cost
                node.pathcost = pathcost
                hinge_hour0_edges.append(DEdge(self.g.initial_node, node, hour0_cost, hour0_kwh))
        # Find the PQ pairs
        self.pq_pairs: List[PriceQuantityUnitless] = []
        forecasted_price_usd_mwh = self.g.params.elec_price_forecast[0] * 10
        # For every possible price
        min_elec_ctskwh, max_elec_ctskwh = -10, 200
        for elec_price_usd_mwh in sorted(list(range(min_elec_ctskwh*10, max_elec_ctskwh*10))+[forecasted_price_usd_mwh]):
            # Update the fake cost of initial node edges with the selected price
            for edge in hinge_hour0_edges:
                if edge.cost >= 1e4: # penalized node
                    edge.fake_cost = edge.cost
                elif edge.rswt_minus_edge_elec is not None: # penalized node
                    edge.fake_cost = edge.rswt_minus_edge_elec * elec_price_usd_mwh/1000
                else:
                    cop = self.g.params.COP(oat=self.g.params.oat_forecast[0], lwt=edge.head.top_temp)
                    edge.fake_cost = edge.hp_heat_out / cop * elec_price_usd_mwh/1000
            # Find the best edge with the given price
            best_edge: DEdge = min(hinge_hour0_edges, key=lambda e: e.head.pathcost + e.fake_cost)
            if best_edge.hp_heat_out < 0: 
                best_edge_neg = max([e for e in hinge_hour0_edges if e.hp_heat_out<0], key=lambda e: e.hp_heat_out)
                best_edge_pos = min([e for e in hinge_hour0_edges if e.hp_heat_out>=0], key=lambda e: e.hp_heat_out)
                best_edge = best_edge_pos if (-best_edge_neg.hp_heat_out >= best_edge_pos.hp_heat_out) else best_edge_neg
            # Find the associated quantity
            cop = self.g.params.COP(oat=self.g.params.oat_forecast[0], lwt=best_edge.head.top_temp)
            best_quantity_kwh = best_edge.hp_heat_out / cop
            best_quantity_kwh = 0 if best_quantity_kwh<0 else best_quantity_kwh
            if not self.pq_pairs:
                self.pq_pairs.append(
                    PriceQuantityUnitless(
                        PriceTimes1000 = int(elec_price_usd_mwh * 1000),
                        QuantityTimes1000 = int(best_quantity_kwh * 1000))
                )
            else:
                # Record a new pair if at least 0.01 kWh of difference in quantity with the previous one
                if self.pq_pairs[-1].QuantityTimes1000 - int(best_quantity_kwh * 1000) > 10:
                    self.pq_pairs.append(
                        PriceQuantityUnitless(
                            PriceTimes1000 = int(elec_price_usd_mwh * 1000),
                            QuantityTimes1000 = int(best_quantity_kwh * 1000))
                    )
        return self.pq_pairs
    
    def to_hingenode(self, node: Union[DNode, HingeNode], time_slice=None):
        return HingeNode(
            time_slice = node.time_slice if time_slice is None else time_slice,
            top_temp = node.top_temp,
            middle_temp = node.middle_temp,
            bottom_temp = node.bottom_temp,
            thermocline1 = node.thermocline1,
            thermocline2 = node.thermocline2,
            params = node.params
        )
    
    def plot_hinge(self):
        sp_time = list(range(self.hinge_hours+2))
        sp_top_temp = []
        sp_middle_temp = []
        sp_bottom_temp = []
        sp_thermocline = []
        sp_thermocline2 = []
        for step in self.hinge_steps:
            sp_top_temp.append(step.top_temp)
            sp_bottom_temp.append(step.bottom_temp)
            sp_thermocline.append(step.thermocline1)
            if step.middle_temp is not None:
                sp_middle_temp.append(step.middle_temp)
                sp_thermocline2.append(step.thermocline2)
            else:
                sp_middle_temp.append(step.bottom_temp)
                sp_thermocline2.append(step.thermocline1)
        norm = Normalize(vmin=self.g.params.available_top_temps[0]-20, vmax=self.g.params.available_top_temps[-1]+20)
        cmap = matplotlib.colormaps['Reds']
        tank_top_colors = [cmap(norm(x)) for x in sp_top_temp]
        tank_middle_colors = [cmap(norm(x)) for x in sp_middle_temp]
        tank_bottom_colors = [cmap(norm(x)) for x in sp_bottom_temp]
        sp_thermocline_reversed1 = [self.g.params.num_layers - x + 1 for x in sp_thermocline]
        sp_thermocline_reversed2 = [self.g.params.num_layers - x + 1 for x in sp_thermocline2]
        bars_top = plt.bar(sp_time, sp_thermocline, bottom=sp_thermocline_reversed1, color=tank_top_colors, alpha=0.7, width=0.9)
        bars_middle = plt.bar(sp_time, [y-x for x,y in zip(sp_thermocline, sp_thermocline2)], 
                              bottom=sp_thermocline_reversed2, color=tank_middle_colors, alpha=0.7, width=0.9)
        bars_bottom = plt.bar(sp_time, sp_thermocline_reversed2, bottom=0, color=tank_bottom_colors, alpha=0.7, width=0.9)
        plt.xlabel('Time [hours]')
        plt.ylabel('Storage state')
        plt.xticks(list(range(self.hinge_hours+2)), labels=list(range(self.hinge_hours+1))+['knit'])
        plt.ylim([0, self.g.params.num_layers])
        plt.yticks([])
        for i, bar in enumerate(bars_top):
            plt.text(bar.get_x() + bar.get_width() / 2, bar.get_y() + bar.get_height() / 2, 
                    f'{int(sp_top_temp[i])}', ha='center', va='center', color='white')
        for i, bar in enumerate(bars_middle):
            if sp_middle_temp[i] != sp_bottom_temp[i]:
                plt.text(bar.get_x() + bar.get_width() / 2, bar.get_y() + bar.get_height() / 2, 
                        f'{int(sp_middle_temp[i])}', ha='center', va='center', color='white')
        for i, bar in enumerate(bars_bottom):
            plt.text(bar.get_x() + bar.get_width() / 2, bar.get_y() + bar.get_height() / 2, 
                    f'{int(sp_bottom_temp[i])}', ha='center', va='center', color='white')
        kwh_el = f"{tuple([round(x/self.g.params.COP(self.g.params.oat_forecast[0]),2) for x in self.best_combination])} kWh_el"
        plt.title(f"{self.best_combination} kWh_th\n{kwh_el}")
        plt.savefig('plot_hinge.png', dpi=130)
        plt.close()

    def quick_plot(self, show=False):
        # Walk along the shortest path (sp) starting from knit
        sp_top_temp = []
        sp_middle_temp = []
        sp_bottom_temp = []
        sp_thermocline = []
        sp_thermocline2 = []
        sp_hp_heat_out = []
        sp_stored_energy = []
        node_i: DNode = self.feasible_branches[self.best_combination]['knitted_to']
        the_end = False
        while not the_end:
            if node_i.next_node is None:
                the_end = True
                sp_hp_heat_out.append(edge_i.hp_heat_out)
            else:
                edge_i = [e for e in self.g.edges[node_i] if e.head==node_i.next_node][0]
                sp_hp_heat_out.append(edge_i.hp_heat_out)
            sp_top_temp.append(node_i.top_temp)
            sp_bottom_temp.append(node_i.bottom_temp)
            sp_thermocline.append(node_i.thermocline1)
            if node_i.middle_temp is not None:
                sp_middle_temp.append(node_i.middle_temp)
                sp_thermocline2.append(node_i.thermocline2)
            else:
                sp_middle_temp.append(node_i.bottom_temp)
                sp_thermocline2.append(node_i.thermocline1)
            sp_stored_energy.append(node_i.energy)
            node_i = node_i.next_node
        sp_soc = [(x-self.g.bottom_node.energy) / (self.g.top_node.energy-self.g.bottom_node.energy) * 100 
                    for x in sp_stored_energy]
        sp_time = list(range(self.g.params.horizon+1))
        start_time = datetime.fromtimestamp(self.g.params.start_time, tz=pytz.timezone("America/New_York"))+timedelta(hours=self.hinge_hours)
        sp_time = [(start_time+timedelta(hours=x)) for x in range(len(sp_time))]

        # Add the discharging part and branching
        sp_time = [
            datetime.fromtimestamp(self.g.params.start_time, tz=pytz.timezone("America/New_York"))
            + timedelta(hours=x) for x in range(self.hinge_hours)
            ] + sp_time
        sp_hp_heat_out = list(self.best_combination) + sp_hp_heat_out
        sp_soc = [(x.energy-self.g.bottom_node.energy) / (self.g.top_node.energy-self.g.bottom_node.energy) * 100 
                    for x in self.hinge_steps[:-2]] + sp_soc
        
        # Plot the shortest path
        fig, ax = plt.subplots(2,1, figsize=(12,6), gridspec_kw={'height_ratios':[8,6]})
        plt.subplots_adjust(hspace=0.3) 
        start = datetime.fromtimestamp(self.g.params.start_time, tz=pytz.timezone("America/New_York")).strftime('%Y-%m-%d %H:%M')

        # Top plot
        self.plot_time = sp_time[:48]
        self.plot_hp = sp_hp_heat_out[:48]
        self.plot_lmp = self.g.params.lmp_forecast[:48]
        self.plot_energy = [x.energy for x in self.hinge_steps[:-2]] + sp_stored_energy
        plot_hours = 12
        ax[0].step(sp_time[:plot_hours], sp_hp_heat_out[:plot_hours], where='post', color='tab:red', alpha=0.6, label='HP')
        ax[0].step(sp_time[:plot_hours], self.g.params.load_forecast[:plot_hours], where='post', color='black', linestyle='dashed', alpha=0.4, label='Load')
        ax[0].legend(loc='upper left')
        ax[0].set_title(f'{start}', fontsize=10)
        ax[0].set_ylabel('Heat [kWh]')
        ax[0].set_ylim([-0.5, 1.5*max(sp_hp_heat_out)])
        ax2 = ax[0].twinx()
        ax2.step(sp_time[:plot_hours], self.g.params.lmp_forecast[:plot_hours], where='post', color='tab:green', alpha=0.8, label='LMP')
        
        ax2.set_ylabel('Electricity price [cts/kWh]')
        yticks = list(set([int(x) for x in self.g.params.lmp_forecast[:plot_hours]]))
        yticks = sorted(yticks+[x+0.5 for x in yticks])
        if len(ax2.get_yticks())>=6 and len(yticks)<=6:
            ax2.set_yticks(yticks)
        ax[0].set_xticks([x for x in sp_time][:plot_hours])
        ax[0].set_xticklabels([f'{x.hour}:00' for x in sp_time][:plot_hours])

        # Bottom plot
        ax[1].plot(sp_time[:plot_hours], sp_soc[:plot_hours], color='black', alpha=0.4, label='SoC')
        ax[1].set_ylabel('Energy in the store [kWh]')
        # ax[1].set_ylim([max(-1,min(sp_soc[:plot_hours])-10),101])
        ax[1].set_yticks([])
        ax[1].set_xticks([x for x in sp_time][:plot_hours])
        ax[1].set_xticklabels([f'{x.hour}:00' for x in sp_time][:plot_hours])

        done_mornings = {}
        done_afternoons = {}
        not_labeled = True
        for i, x in enumerate(sp_time[:plot_hours]):
            if x.hour in [7,8,9,10,11,16,17,18,19] and x.weekday() not in [5,6]:
                if x.hour in [7,8,9,10,11] and x.date() not in done_mornings:
                    end_index = i+5-(x.hour-7) if i==0 else min(i+5, plot_hours-1)
                    done_mornings[x.date()] = True
                    if not_labeled:
                        ax2.axvspan(sp_time[i], sp_time[end_index], color='tab:green', alpha=0.05, label='Onpeak')
                        not_labeled = False
                    else:
                        ax2.axvspan(sp_time[i], sp_time[end_index], color='tab:green', alpha=0.05)
                elif x.hour in [16,17,18,19] and x.date() not in done_afternoons:
                    end_index = i+4-(x.hour-16) if i==0 else min(i+4, plot_hours-1)
                    done_afternoons[x.date()] = True
                    if not_labeled:
                        ax2.axvspan(sp_time[i], sp_time[end_index], color='tab:green', alpha=0.05, label='Onpeak')
                        not_labeled = False
                    else:
                        ax2.axvspan(sp_time[i], sp_time[end_index], color='tab:green', alpha=0.05)
        ax2.legend(loc='upper right')

        # plt.tight_layout()
        plt.savefig('plot_quick.png', dpi=100, bbox_inches='tight')
        if show:
            plt.show()
        plt.close()
    
    def export_to_excel(self):
        # Sort nodes by energy and assign an index
        for time_slice in range(self.g.params.horizon+1):
            self.g.nodes_by_energy = sorted(self.g.nodes[time_slice], key=lambda x: (x.energy, x.top_temp), reverse=True)
            for n in self.g.nodes[time_slice]:
                n.index = self.g.nodes_by_energy.index(n)+1

        # Along the shortest path
        electricitiy_used, heat_delivered = [], []
        node_i = self.g.initial_node
        while node_i.next_node is not None:
            losses = self.g.params.storage_losses_percent/100 * (node_i.energy-self.g.bottom_node.energy)
            if self.g.params.load_forecast[node_i.time_slice]==0 and losses>0 and losses<self.g.params.energy_between_nodes[node_i.top_temp]:
                losses = self.g.params.energy_between_nodes[node_i.top_temp] + 1/1e9
            store_heat_in = node_i.next_node.energy - node_i.energy
            hp_heat_out = store_heat_in + self.g.params.load_forecast[node_i.time_slice] + losses
            cop = self.g.params.COP(oat=self.g.params.oat_forecast[node_i.time_slice], lwt=node_i.next_node.top_temp)
            heat_delivered.append(hp_heat_out)
            electricitiy_used.append(hp_heat_out/cop)
            node_i = node_i.next_node
        
        # First dataframe: the Dijkstra graph
        dijkstra_pathcosts = {}
        dijkstra_pathcosts['Model'] = [repr(x) for x in self.g.nodes_by_energy]
        dijkstra_pathcosts['Energy (relative)'] = [round(x.energy-self.g.bottom_node.energy,2) for x in self.g.nodes_by_energy]
        dijkstra_pathcosts['Index'] = list(range(1,len(self.g.nodes_by_energy)+1))
        dijkstra_nextnodes = dijkstra_pathcosts.copy()
        for h in range(self.g.params.horizon):
            dijkstra_pathcosts[h] = [round(x.pathcost,2) for x in sorted(self.g.nodes[h], key=lambda x: x.index)]
            dijkstra_nextnodes[h] = [x.next_node.index for x in sorted(self.g.nodes[h], key=lambda x: x.index)]
        dijkstra_pathcosts[self.g.params.horizon] = [0 for x in self.g.nodes[self.g.params.horizon]]
        dijkstra_nextnodes[self.g.params.horizon] = [np.nan for x in self.g.nodes[self.g.params.horizon]]
        dijkstra_pathcosts_df = pd.DataFrame(dijkstra_pathcosts)
        dijkstra_nextnodes_df = pd.DataFrame(dijkstra_nextnodes)
        
        # Second dataframe: the forecasts
        start_time = datetime.fromtimestamp(self.g.params.start_time, tz=pytz.timezone("America/New_York"))
        forecast_df = pd.DataFrame({'Forecast':['0'], 'Unit':['0'], **{h: [0.0] for h in range(self.g.params.horizon)}})
        forecast_df.loc[0] = ['Hour'] + [start_time.strftime("%d/%m/%Y")] + [(start_time + timedelta(hours=x)).hour for x in range(self.g.params.horizon)]
        forecast_df.loc[1] = ['Price - total'] + ['cts/kWh'] + self.g.params.elec_price_forecast
        forecast_df.loc[2] = ['Price - distribution'] + ['cts/kWh'] + self.g.params.dist_forecast
        forecast_df.loc[3] = ['Price - LMP'] + ['cts/kWh'] + self.g.params.lmp_forecast
        forecast_df.loc[4] = ['Heating load'] + ['kW'] + [round(x,2) for x in self.g.params.load_forecast]
        forecast_df.loc[5] = ['OAT'] + ['F'] + [round(x,2) for x in self.g.params.oat_forecast]
        forecast_df.loc[6] = ['Required SWT'] + ['F'] + [round(x) for x in self.g.params.rswt_forecast]
        
        # Third dataframe: the shortest path
        shortestpath_df = pd.DataFrame({'Shortest path':['0'], 'Unit':['0'], **{h: [0.0] for h in range(self.g.params.horizon+1)}})
        shortestpath_df.loc[0] = ['Electricity used'] + ['kWh'] + [round(x,3) for x in electricitiy_used] + [0]
        shortestpath_df.loc[1] = ['Heat delivered'] + ['kWh'] + [round(x,3) for x in heat_delivered] + [0]
        shortestpath_df.loc[2] = ['Cost - total'] + ['cts'] + [round(x*y,2) for x,y in zip(electricitiy_used, self.g.params.elec_price_forecast)] + [0]
        shortestpath_df.loc[3] = ['Cost - distribution'] + ['cts'] + [round(x*y,2) for x,y in zip(electricitiy_used, self.g.params.dist_forecast)] + [0]
        shortestpath_df.loc[4] = ['Cost - LMP'] + ['cts'] + [round(x*y,2) for x,y in zip(electricitiy_used, self.g.params.lmp_forecast)] + [0]
        
        # Fourth dataframe: the results
        total_usd = round(self.g.initial_node.pathcost,2)
        total_elec = round(sum(electricitiy_used),2)
        total_heat = round(sum(heat_delivered),2)
        next_index = self.g.initial_node.next_node.index
        results = ['Cost ($)', total_usd, 'Electricity (kWh)', total_elec, 'Heat (kWh)', total_heat, 'Next step index', next_index]
        results_df = pd.DataFrame({'RESULTS':results})
        
        # Highlight shortest path
        highlight_positions = []
        node_i = self.g.initial_node
        while node_i.next_node is not None:
            highlight_positions.append((node_i.index+len(forecast_df)+len(shortestpath_df)+2, 3+node_i.time_slice))
            node_i = node_i.next_node
        highlight_positions.append((node_i.index+len(forecast_df)+len(shortestpath_df)+2, 3+node_i.time_slice))
        
        # Add the parameters to a seperate sheet
        parameters = self.g.params.config.to_dict()
        parameters_df = pd.DataFrame(list(parameters.items()), columns=['Variable', 'Value'])

        # Add the PQ pairs to a seperate sheet and plot the curve
        pq_pairs = self.generate_bid()
        prices = [x.PriceTimes1000 for x in pq_pairs]
        quantities = [x.QuantityTimes1000/1000 for x in pq_pairs]
        pqpairs_df = pd.DataFrame({'price':[x/1000 for x in prices], 'quantity':quantities})
        # To plot quantities on x-axis and prices on y-axis
        ps, qs = [], []
        index_p = 0
        expected_price_usd_mwh = self.g.params.elec_price_forecast[0] * 10
        for p in sorted(list(range(min(prices), max(prices)+1)) + [expected_price_usd_mwh*1000]):
            ps.append(p/1000)
            if index_p+1 < len(prices) and p >= prices[index_p+1]:
                index_p += 1
            if p == expected_price_usd_mwh*1000:
                interesection = (quantities[index_p], expected_price_usd_mwh)
            qs.append(quantities[index_p])
        plt.plot(qs, ps, label='demand (bid)')
        prices = [x.PriceTimes1000/1000 for x in pq_pairs]
        plt.scatter(quantities, prices)
        plt.plot([min(quantities)-1, max(quantities)+1],[expected_price_usd_mwh]*2, label="supply (expected market price)")
        plt.scatter(interesection[0], interesection[1])
        plt.text(interesection[0]+0.25, interesection[1]+15, f'({round(interesection[0],3)}, {round(interesection[1],1)})', fontsize=10, color='tab:orange')
        plt.xticks(quantities)
        if min([abs(x-expected_price_usd_mwh) for x in prices]) < 5:
            plt.yticks(prices)
        else:
            plt.yticks(prices + [expected_price_usd_mwh])
        plt.ylabel("Price [USD/MWh]")
        plt.xlabel("Quantity [kWh]")
        plt.grid(alpha=0.3)
        plt.legend()
        plt.tight_layout()
        plt.savefig('plot_pq.png', dpi=130)
        plt.close()

        # Write to Excel
        os.makedirs('results', exist_ok=True)
        start = datetime.fromtimestamp(self.g.params.start_time, tz=pytz.timezone("America/New_York")).strftime('%Y-%m-%d %H:%M')
        # file_path = os.path.join('results', f'result_{start}.xlsx')
        file_path = 'result.xlsx'
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:

            # Add summary plot
            quick_plot_sheet = writer.book.create_sheet(title='Quick look')
            quick_plot_sheet.add_image(Image('plot_quick.png'), 'A1')

            results_df.to_excel(writer, index=False, sheet_name='Pathcost')
            results_df.to_excel(writer, index=False, sheet_name='Next node')
            forecast_df.to_excel(writer, index=False, startcol=1, sheet_name='Pathcost')
            forecast_df.to_excel(writer, index=False, startcol=1, sheet_name='Next node')
            shortestpath_df.to_excel(writer, index=False, startcol=1, startrow=len(forecast_df)+1, sheet_name='Pathcost')
            shortestpath_df.to_excel(writer, index=False, startcol=1, startrow=len(forecast_df)+1, sheet_name='Next node')
            dijkstra_pathcosts_df.to_excel(writer, index=False, startrow=len(forecast_df)+len(shortestpath_df)+2, sheet_name='Pathcost')
            dijkstra_nextnodes_df.to_excel(writer, index=False, startrow=len(forecast_df)+len(shortestpath_df)+2, sheet_name='Next node')
            parameters_df.to_excel(writer, index=False, sheet_name='Parameters')
            
            # Add plot in a seperate sheet
            self.g.plot(show=False)
            plot_sheet = writer.book.create_sheet(title='Plot')
            plot_sheet.add_image(Image('plot.png'), 'A1')
            plot_sheet.add_image(Image('plot_hinge.png'), 'T1')

            # Add plot in a seperate sheet
            plot2_sheet = writer.book.create_sheet(title='PQ pairs')
            pqpairs_df.to_excel(writer, index=False, sheet_name='PQ pairs')
            plot2_sheet.add_image(Image('plot_pq.png'), 'C1')

            # Layout
            pathcost_sheet = writer.sheets['Pathcost']
            nextnode_sheet = writer.sheets['Next node']
            parameters_sheet = writer.sheets['Parameters']
            for row in pathcost_sheet['A1:A10']:
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
            for row in nextnode_sheet['A1:A10']:
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
            for row in parameters_sheet[f'B1:B{len(parameters_df)+1}']:
                for cell in row:
                    cell.alignment = Alignment(horizontal='right')
            pathcost_sheet.column_dimensions['A'].width = 17.5
            pathcost_sheet.column_dimensions['B'].width = 15
            pathcost_sheet.column_dimensions['C'].width = 15
            nextnode_sheet.column_dimensions['A'].width = 17.5
            nextnode_sheet.column_dimensions['B'].width = 15
            nextnode_sheet.column_dimensions['C'].width = 15
            parameters_sheet.column_dimensions['A'].width = 40
            parameters_sheet.column_dimensions['B'].width = 70
            pathcost_sheet.freeze_panes = 'D16'
            nextnode_sheet.freeze_panes = 'D16'

            # Highlight shortest path
            highlight_fill = PatternFill(start_color='72ba93', end_color='72ba93', fill_type='solid')
            for row in range(len(forecast_df)+len(shortestpath_df)+2):
                pathcost_sheet.cell(row=row+1, column=1).fill = highlight_fill
                nextnode_sheet.cell(row=row+1, column=1).fill = highlight_fill
            for row, col in highlight_positions:
                pathcost_sheet.cell(row=row+1, column=col+1).fill = highlight_fill
                nextnode_sheet.cell(row=row+1, column=col+1).fill = highlight_fill

        os.remove('plot.png')        
        os.remove('plot_pq.png')
        os.remove('plot_quick.png')
        os.remove('plot_hinge.png')


if __name__ == '__main__':

    # ----------------------------------
    # Load a FLO params
    def from_dict_msg(data):
        message = MessageSql(
                message_id=data["MessageId"],
                from_alias=data["FromAlias"],
                message_type_name=data["MessageTypeName"],
                message_persisted_ms=data["MessagePersistedMs"],
                payload=data["Payload"],
                message_created_ms=data.get("MessageCreatedMs")  # This is optional
            )
        return message
    with open('messages.json', 'r') as file:
        messages_dict = json.load(file)
    message_loaded = from_dict_msg(messages_dict)
    flo_params = FloParamsHouse0(**message_loaded.payload)
    # ----------------------------------
    
    f = FloHinge(flo_params, hinge_hours=5, num_nodes=[10,3,3,3,3])
    f.generate_bid()
    f.export_to_excel()