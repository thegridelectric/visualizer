from datetime import datetime, timedelta
import numpy as np
from typing import Dict, List, Tuple
import os
import pytz
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.dates import DateFormatter
from matplotlib.colors import Normalize, ListedColormap, BoundaryNorm
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.drawing.image import Image
from named_types import PriceQuantityUnitless, FloParamsHouse0
from typing import Optional
import time

MIN_DIFFERENCE_F = 10
STEP_F = 10
NUM_LAYERS = 12

def to_kelvin(t):
    return (t-32)*5/9 + 273.15

def to_celcius(t):
    return (t-32)*5/9

class DParams():
    def __init__(self, config: FloParamsHouse0) -> None:
        self.config = config
        self.start_time = config.StartUnixS
        self.horizon = config.HorizonHours
        self.num_layers = min(config.NumLayers, NUM_LAYERS)
        self.storage_volume = config.StorageVolumeGallons
        self.max_hp_elec_in = config.HpMaxElecKw
        self.min_hp_elec_in = config.HpMinElecKw
        self.initial_top_temp = config.InitialTopTempF
        self.initial_bottom_temp = config.InitialBottomTempF
        self.initial_thermocline = min(config.InitialThermocline, self.num_layers)
        self.storage_losses_percent = config.StorageLossesPercent
        self.reg_forecast = [x/10 for x in config.RegPriceForecast[:self.horizon]]
        self.dist_forecast = [x/10 for x in config.DistPriceForecast[:self.horizon]]
        self.lmp_forecast = [x/10 for x in config.LmpForecast[:self.horizon]]
        self.elec_price_forecast = [rp+dp+lmp for rp,dp,lmp in zip(self.reg_forecast, self.dist_forecast, self.lmp_forecast)]
        self.oat_forecast = config.OatForecastF[:self.horizon]
        self.ws_forecast = config.WindSpeedForecastMph[:self.horizon]
        self.alpha = config.AlphaTimes10/10
        self.beta = config.BetaTimes100/100
        self.gamma = config.GammaEx6/1e6
        self.no_power_rswt = -self.alpha/self.beta
        self.intermediate_power = config.IntermediatePowerKw
        self.intermediate_rswt = config.IntermediateRswtF
        self.dd_power = config.DdPowerKw
        self.dd_rswt = config.DdRswtF
        self.dd_delta_t = config.DdDeltaTF
        self.hp_is_off = config.HpIsOff
        self.hp_turn_on_minutes = config.HpTurnOnMinutes
        self.quadratic_coefficients = self.get_quadratic_coeffs()
        self.load_forecast = [self.required_heating_power(oat,ws) for oat,ws in zip(self.oat_forecast,self.ws_forecast)]
        self.rswt_forecast = [self.required_swt(x) for x in self.load_forecast]
        # Modify load forecast to include energy available in the buffer
        available_buffer = config.BufferAvailableKwh
        i = 0
        while available_buffer > 0 and i < len(self.load_forecast):
            load_backup = self.load_forecast[i]
            self.load_forecast[i] = self.load_forecast[i] - min(available_buffer, self.load_forecast[i])
            available_buffer = available_buffer - min(available_buffer, load_backup)
            i += 1
        # Modify load forecast to include energy available in the house (zones above thermostat)
        available_house = config.HouseAvailableKwh
        i = 0
        if available_house < 0:
            self.load_forecast[0] += -available_house
        else:
            while available_house > 0 and i < len(self.load_forecast):
                load_backup = self.load_forecast[i]
                self.load_forecast[i] = self.load_forecast[i] - min(available_house, self.load_forecast[i])
                available_house = available_house - min(available_house, load_backup)
                i += 1
        self.check_hp_sizing()
        # TODO: add to config
        self.min_cop = 1
        self.max_cop = 3
        self.soft_constraint: bool = True
        # First time step can be shorter than an hour
        if datetime.fromtimestamp(self.start_time).minute > 0:
            self.fraction_of_hour_remaining: float = datetime.fromtimestamp(self.start_time).minute / 60
        else:
            self.fraction_of_hour_remaining: float = 1
        self.load_forecast[0] = self.load_forecast[0]*self.fraction_of_hour_remaining
        
    def check_hp_sizing(self):
        max_load_elec = max(self.load_forecast) / self.COP(min(self.oat_forecast), max(self.rswt_forecast))
        if max_load_elec > self.max_hp_elec_in:
            error_text = f"\nThe current parameters indicate that on the coldest hour of the forecast ({min(self.oat_forecast)} F):"
            error_text += f"\n- The heating requirement is {round(max(self.load_forecast),2)} kW"
            error_text += f"\n- The COP is {round(self.COP(min(self.oat_forecast), max(self.rswt_forecast)),2)}"
            error_text += f"\n=> Need a HP that can reach {round(max_load_elec,2)} kW electrical power"
            error_text += f"\n=> The given HP is undersized ({self.max_hp_elec_in} kW electrical power)"
            print(error_text)
        
    def COP(self, oat, lwt=None):
        if oat < self.config.CopMinOatF: 
            return self.config.CopMin
        else:
            return self.config.CopIntercept + self.config.CopOatCoeff * oat

    def required_heating_power(self, oat, ws):
        r = self.alpha + self.beta*oat + self.gamma*ws
        return r if r>0 else 0

    def delivered_heating_power(self, swt):
        a, b, c = self.quadratic_coefficients
        d = a*swt**2 + b*swt + c
        return d if d>0 else 0

    def required_swt(self, rhp):
        a, b, c = self.quadratic_coefficients
        c2 = c - rhp
        return (-b + (b**2-4*a*c2)**0.5)/(2*a)

    def delta_T(self, swt):
        return 20
    
    def delta_T_inverse(self, rwt: float) -> float:
        a, b, c = self.quadratic_coefficients
        aa = -self.dd_delta_t/self.dd_power * a
        bb = 1-self.dd_delta_t/self.dd_power * b
        cc = -self.dd_delta_t/self.dd_power * c - rwt
        if bb**2-4*aa*cc < 0 or (-bb + (bb**2-4*aa*cc)**0.5)/(2*aa) - rwt > 30:
            return 30
        return (-bb + (bb**2-4*aa*cc)**0.5)/(2*aa) - rwt
    
    def get_quadratic_coeffs(self):
        x_rswt = np.array([self.no_power_rswt, self.intermediate_rswt, self.dd_rswt])
        y_hpower = np.array([0, self.intermediate_power, self.dd_power])
        A = np.vstack([x_rswt**2, x_rswt, np.ones_like(x_rswt)]).T
        return [float(x) for x in np.linalg.solve(A, y_hpower)] 


class DNode():
    def __init__(self, 
                 parameters: DParams,
                 time_slice:int, 
                 top_temp:float, 
                 thermocline1:int, 
                 bottom_temp:float, 
                 thermocline2: Optional[int]=None, 
                 middle_temp: Optional[float]=None
                 ):
        self.params = parameters
        # State
        if not middle_temp:
            middle_temp = top_temp-MIN_DIFFERENCE_F if top_temp>=100+2*MIN_DIFFERENCE_F else 100
            thermocline2 = thermocline1
        self.time_slice = time_slice
        self.top_temp = top_temp
        self.middle_temp = middle_temp
        self.bottom_temp = bottom_temp
        self.thermocline1 = thermocline1
        self.thermocline2 = thermocline2
        # Dijkstra's algorithm
        self.pathcost = 0 if time_slice==parameters.horizon else 1e9
        self.next_node = None
        # Absolute energy level
        self.energy = self.get_energy()
        self.index = None

    def __repr__(self):
        return f"[{self.time_slice}]{self.top_temp}({self.thermocline1}){self.middle_temp}({self.thermocline2}){self.bottom_temp}"
        
    def get_energy(self):
        m_layer_kg = self.params.storage_volume*3.785 / self.params.num_layers
        kWh_top = self.thermocline1*m_layer_kg * 4.187/3600 * to_kelvin(self.top_temp)
        kWh_midlle = (self.thermocline2-self.thermocline1)*m_layer_kg * 4.187/3600 * to_kelvin(self.middle_temp)
        kWh_bottom = (self.params.num_layers-self.thermocline2)*m_layer_kg * 4.187/3600 * to_kelvin(self.bottom_temp)
        return kWh_top + kWh_midlle + kWh_bottom
    
    def plot(self):
        norm = Normalize(vmin=100-20, vmax=170+20)
        cmap = matplotlib.colormaps['Reds'] 
        tank_top_colors = [cmap(norm(x)) for x in [self.top_temp]]
        tank_middle_colors = [cmap(norm(x)) for x in [self.middle_temp]]
        tank_bottom_colors = [cmap(norm(x)) for x in [self.bottom_temp]]
        thermocline1_reversed = self.params.num_layers - self.thermocline1
        thermocline2_reversed = self.params.num_layers - self.thermocline2
        bars_top = plt.bar([0], 
                           [self.thermocline1], 
                           bottom=thermocline1_reversed, 
                           color=tank_top_colors, alpha=0.9, width=0.5)
        bars_middle = plt.bar([0], 
                              [self.thermocline2 - self.thermocline1], 
                              bottom=[thermocline2_reversed], 
                              color=tank_middle_colors, alpha=0.9, width=0.5)
        bars_bottom = plt.bar([0], 
                              [thermocline2_reversed], 
                              bottom=[0], 
                              color=tank_bottom_colors, alpha=0.9, width=0.5)
        plt.xlim([-1,1])
        plt.xticks([])
        for bar in bars_top:
            plt.text(bar.get_x() + bar.get_width()/2, bar.get_y() + bar.get_height()/2, 
                     f'{int(self.top_temp)}', ha='center', va='center', color='white')
        for bar in bars_middle:
            plt.text(bar.get_x() + bar.get_width()/2, bar.get_y() + bar.get_height()/2, 
                    f'{int(self.middle_temp)}', ha='center', va='center', color='white')
        for bar in bars_bottom:
            plt.text(bar.get_x() + bar.get_width()/2, bar.get_y() + bar.get_height()/2, 
                     f'{int(self.bottom_temp)}', ha='center', va='center', color='white')
        plt.title(repr(self))
        plt.show()


class DEdge():
    def __init__(self, tail:DNode, head:DNode, cost:float, hp_heat_out:float):
        self.tail: DNode = tail
        self.head: DNode = head
        self.cost = cost
        self.hp_heat_out = hp_heat_out
        self.fake_cost: Optional[float] = None

    def __repr__(self):
        return f"Edge[{self.tail} --cost:{round(self.cost,3)}, hp:{round(self.hp_heat_out,2)}--> {self.head}]"
    

class DGraph():
    def __init__(self, config: FloParamsHouse0):
        self.params = DParams(config)
        self.nodes: Dict[int, List[DNode]] = {h: [] for h in range(self.params.horizon+1)}
        self.edges: Dict[DNode, List[DEdge]] = {}
        self.time_spent_in_charge = 0
        self.time_spent_in_discharge = 0
        st = time.time()
        self.create_nodes()
        self.create_edges()
        self.time_spent_in_total = time.time()-st

    def create_nodes(self):
        if STEP_F == 5:
            self.top_temps = sorted(list(range(110,170+2*STEP_F,STEP_F)), reverse=True) # 175, 170, ..., 110
        else:
            self.top_temps = sorted(list(range(110,170+STEP_F,STEP_F)), reverse=True) # 175, 170, ..., 110
        self.middle_temps = [x-MIN_DIFFERENCE_F for x in (self.top_temps[:-1] if STEP_F==10 else self.top_temps[:-2])] 
        self.bottom_temps = [100]
        # Initial node temperatures
        initial_top_temp = min(self.top_temps+[100,80], key=lambda x: abs(x-self.params.initial_top_temp))
        initial_middle_temp = min(self.middle_temps, key=lambda x: abs(x-self.params.initial_bottom_temp))
        initial_bottom_temp = min(self.bottom_temps, key=lambda x: abs(x-self.params.initial_bottom_temp))
        # If initial node is a cold node
        if initial_top_temp == 110:
            initial_middle_temp = 100
        if initial_top_temp in [100,80]:
            initial_middle_temp = initial_top_temp-20
            initial_bottom_temp = initial_top_temp-20
        # Initial node
        self.initial_node = DNode(
            parameters=self.params,
            time_slice=0,
            top_temp=initial_top_temp,
            middle_temp=initial_middle_temp,
            bottom_temp=initial_bottom_temp,
            thermocline1=self.params.initial_thermocline,
            thermocline2=self.params.num_layers if initial_top_temp not in [110,100,80] else self.params.initial_thermocline,
        )
        self.nodes[0] = [self.initial_node]
        print(f"Initial node: {self.initial_node}")

        self.storage_full = False
        if STEP_F == 5 and self.params.initial_top_temp > max(self.top_temps)-STEP_F:
            print("Storage is currently full, don't allow any node with more energy than this")
            self.storage_full = True

        thermocline_combinations = []
        for t1 in range(1,self.params.num_layers+1):
            for t2 in range(1,self.params.num_layers+1):
                if t2>=t1:
                    thermocline_combinations.append((t1,t2))
        print(f"=> {len(thermocline_combinations)} thermocline combinations")

        temperature_combinations = []
        for t in self.top_temps:
            for m in self.middle_temps:
                for b in self.bottom_temps:
                    if b<=m-MIN_DIFFERENCE_F and m<=t-MIN_DIFFERENCE_F:
                        temperature_combinations.append((t,m,b))
        print(f"=> {len(temperature_combinations)} temperature combinations")

        # Add cases where we can't have t >= m+10 >= b+10
        if STEP_F==5: temperature_combinations += [(115,100,100)]
        temperature_combinations += [(110,100,100)]

        # Add colder temperatures
        temperature_combinations += [(100,80,80), (80,60,60)]
        
        total_nodes=0
        for tmb in temperature_combinations:
            for h in range(self.params.horizon+1):
                for th in thermocline_combinations:
                    t, m, b = tmb
                    th1, th2 = th
                    if m==b and th1!=th2:
                        continue
                    node = DNode(
                        time_slice=h,
                        top_temp=t,
                        middle_temp=m,
                        bottom_temp=b,
                        thermocline1=th1,
                        thermocline2=th2,
                        parameters=self.params
                        )
                    if self.storage_full and node.energy>=self.initial_node.energy:
                        continue
                    self.nodes[h].append(node)
                    total_nodes += 1
                
        print(f"=> Created a total of {int(total_nodes/49)} nodes per layer")
        # for n in self.nodes[0][:30]:
        #     n.plot()

    def create_edges(self):
        print("Creating all edges...")
        
        self.bottom_node_energy = DNode(
            time_slice=0,
            top_temp=80,
            thermocline1=1,
            bottom_temp=70,
            parameters=self.params
        ).energy

        self.top_node_energy = DNode(
            time_slice=0,
            top_temp=175 if STEP_F==5 else 170,
            thermocline1=self.params.num_layers,
            bottom_temp=175 if STEP_F==5 else 170,
            parameters=self.params
        ).energy

        for h in range(self.params.horizon):

            # Find the maximum heat the HP can put out
            max_hp_elec_in = self.params.max_hp_elec_in
            if h==0:
                max_hp_elec_in = max_hp_elec_in * self.params.fraction_of_hour_remaining
                max_hp_elec_in = (((1-self.params.hp_turn_on_minutes/60) if self.params.hp_is_off else 1) * max_hp_elec_in)
            else:
                # Since we can't know if the HP was on or off after hour 0, remove half of the turn on time
                # Overestimating less when turning on, underestimating a little when already on
                max_hp_elec_in = ((1-self.params.hp_turn_on_minutes/2/60) * max_hp_elec_in)            
            cop = self.params.COP(oat=self.params.oat_forecast[h])
            max_hp_heat_out = max_hp_elec_in * cop
            
            for node_now in self.nodes[h]:
                self.edges[node_now] = []

                losses = self.params.storage_losses_percent/100 * (node_now.energy-self.bottom_node_energy)
                load = self.params.load_forecast[h]
                rswt = self.params.rswt_forecast[h]

                # Ajust the maximum heat output
                store_heat_in_for_full = self.top_node_energy - node_now.energy
                hp_heat_out_for_full = store_heat_in_for_full + load + losses
                if hp_heat_out_for_full < max_hp_heat_out:
                    if hp_heat_out_for_full > 10: # TODO make this a min_hp_heat out parameter
                        allowed_hp_heat_out = [0, hp_heat_out_for_full]
                    else:
                        allowed_hp_heat_out = [0]
                else:
                    allowed_hp_heat_out = [0, max_hp_heat_out]

                for hp_heat_out in allowed_hp_heat_out:
                    store_heat_in = hp_heat_out - load - losses
                    node_next = self.model_accurately(node_now, store_heat_in)
                    cost = self.params.elec_price_forecast[h]/100*hp_heat_out/cop
                    if store_heat_in < 0 and load > 0:
                        if node_now.top_temp<rswt or node_next.top_temp<rswt:
                            cost += 1e5
                    self.edges[node_now].append(DEdge(node_now, node_next, cost, hp_heat_out))

            print(f"Done for hour {h}")

    def model_accurately(self, node_now:DNode, store_heat_in:float, print_detail:bool=False):
        if store_heat_in > 0:
            if print_detail: print(f"Charge {node_now} by {store_heat_in}")
            st = time.time()
            next_node = self.charge(node_now, store_heat_in, print_detail)
            self.time_spent_in_charge += (time.time()-st)
        elif store_heat_in < -1:
            if print_detail: print(f"Discharge {node_now} by {-store_heat_in}")
            st = time.time()
            next_node = self.discharge(node_now, store_heat_in, print_detail)
            self.time_spent_in_discharge += (time.time()-st)
        else:
            if print_detail: print("IDLE")
            next_node = [
                x for x in self.nodes[node_now.time_slice+1]
                if x.params==node_now.params
                and x.thermocline1==node_now.thermocline1
                and x.thermocline2==node_now.thermocline2
                and x.top_temp==node_now.top_temp
                and x.bottom_temp==node_now.bottom_temp
                and x.middle_temp==node_now.middle_temp
                ][0]
        return next_node
        
    def charge(self, n: DNode, store_heat_in: float, print_detail: bool) -> DNode:
        next_node_energy = n.energy + store_heat_in

        # Charging from cold states
        if n.top_temp<=100:
            next_node_top_temp = n.top_temp
            next_node_bottom_temp = n.bottom_temp
            next_node_thermocline = self.find_thermocline(next_node_top_temp, next_node_bottom_temp, next_node_energy)
            if next_node_thermocline > self.params.num_layers and n.top_temp == 80:
                next_node_top_temp = 100
                next_node_bottom_temp = 80
                next_node_thermocline = self.find_thermocline(next_node_top_temp, next_node_bottom_temp, next_node_energy)
            # Need to rise above cold states
            if next_node_thermocline > self.params.num_layers:
                # At this point we know we have
                next_node_top_temp = 110
                next_node_bottom_temp = 100

        else:
            # If there is a bottom layer
            if n.thermocline2 < self.params.num_layers:
                heated_bottom = n.bottom_temp + self.params.delta_T(n.bottom_temp)
                if print_detail: print(f"heated_bottom = {heated_bottom}")

                # Find the new top temperature after mixing (or not)
                if heated_bottom < n.top_temp:
                    top_and_middle_mixed = (n.top_temp*n.thermocline1 + n.middle_temp*(n.thermocline2-n.thermocline1))/n.thermocline2
                    if print_detail: print(f"top_and_middle_mixed = {top_and_middle_mixed}")
                    top_and_middle_and_heated_bottom_mixed = (
                        (top_and_middle_mixed*n.thermocline2 + heated_bottom*(self.params.num_layers-n.thermocline2))/self.params.num_layers
                        )
                    if print_detail: print(f"top_and_middle_and_heated_bottom_mixed = {round(top_and_middle_and_heated_bottom_mixed,1)}")
                    next_node_top_temp = round(top_and_middle_and_heated_bottom_mixed)
                else:
                    next_node_top_temp = heated_bottom   
                
                # Bottom layer stays the same
                next_node_bottom_temp = n.bottom_temp

            # If there is no bottom layer but there is a middle layer
            elif n.thermocline1 < self.params.num_layers:     
                heated_middle = n.middle_temp + self.params.delta_T(n.middle_temp)
                if print_detail: print(f"heated_middle = {heated_middle}")

                # Find the new top temperature after mixing (or not)
                if heated_middle < n.top_temp:
                    top_and_heated_middle_mixed = (
                        (n.top_temp*n.thermocline1 + heated_middle*(self.params.num_layers-n.thermocline1))/self.params.num_layers
                        )
                    if print_detail: print(f"top_and_heated_middle_mixed = {round(top_and_heated_middle_mixed,1)}")
                    next_node_top_temp = round(top_and_heated_middle_mixed)
                else:
                    next_node_top_temp = heated_middle   

                # Bottom layer is the middle
                next_node_bottom_temp = n.middle_temp

            # If there is only a top layer
            else:
                heated_top = n.top_temp + self.params.delta_T(n.top_temp)
                if print_detail: print(f"heated_top = {heated_top}")
                next_node_top_temp = heated_top   
                # Bottom layer is the top
                next_node_bottom_temp = n.top_temp

        # Starting with that top and current bottom, find the thermocline position that matches the next node energy
        next_node_thermocline = self.find_thermocline(next_node_top_temp, next_node_bottom_temp, next_node_energy)
        if print_detail: print(f"Next node ({next_node_top_temp}, {next_node_bottom_temp}) thermocline: {next_node_thermocline}")
        while next_node_thermocline > self.params.num_layers:
            next_node_bottom_temp = next_node_top_temp
            next_node_top_temp = round(next_node_top_temp + self.params.delta_T(next_node_top_temp))
            next_node_thermocline = self.find_thermocline(next_node_top_temp, next_node_bottom_temp, next_node_energy)
            if print_detail: print(f"Next node ({next_node_top_temp}, {next_node_bottom_temp}) thermocline: {next_node_thermocline}")

        if next_node_top_temp <= 100:
            node_next_true = DNode(
                parameters = self.params,
                time_slice = n.time_slice+1,
                top_temp = next_node_top_temp,
                middle_temp = next_node_bottom_temp,
                bottom_temp = next_node_bottom_temp,
                thermocline1 = next_node_thermocline,
                thermocline2 = next_node_thermocline,
            )
        
        else:
            if next_node_bottom_temp not in self.bottom_temps:
                node_next_true = DNode(
                    parameters = self.params,
                    time_slice = n.time_slice+1,
                    top_temp = next_node_top_temp if next_node_thermocline>0 else next_node_bottom_temp,
                    middle_temp = next_node_bottom_temp if next_node_thermocline>0 else None,
                    bottom_temp = n.bottom_temp,
                    thermocline1 = next_node_thermocline if next_node_thermocline>0 else self.params.num_layers,
                    thermocline2 = self.params.num_layers if next_node_thermocline>0 else None,
                )
            else:
                node_next_true = DNode(
                    parameters = self.params,
                    time_slice = n.time_slice+1,
                    top_temp = next_node_top_temp,
                    middle_temp = None,
                    bottom_temp = next_node_bottom_temp,
                    thermocline1 = next_node_thermocline,
                    thermocline2 = None
                )

        node_next = self.find_closest_node(node_next_true, print_detail)
        if print_detail: print(f"True: {node_next_true}, associated to {node_next}")
        return node_next
        
    def find_thermocline(self, top_temp, bottom_temp, energy):
        top, bottom = to_kelvin(top_temp), to_kelvin(bottom_temp)
        m_layer_kg = self.params.storage_volume*3.785 / self.params.num_layers    
        if top==bottom: top+=1  
        return int(1/(top-bottom)*(energy/(m_layer_kg*4.187/3600)-(-0.5*top+(self.params.num_layers+0.5)*bottom)))
    
    def discharge(self, n: DNode, store_heat_in: float, print_detail: bool) -> DNode:
        next_node_energy = n.energy + store_heat_in
        candidate_nodes: List[DNode] = []
        if print_detail: print(f"Current energy {round(n.energy,2)}, looking for {round(next_node_energy,2)}")
        # Starting from current node
        th1 = n.thermocline1-1
        th2 = n.thermocline2-1

        # Node to discharge is a cold node
        if n.top_temp <= 100:
            if n.top_temp==80 and th1==0:
                return min(self.nodes[n.time_slice+1], key=lambda x: x.energy)
            # Go through the top being at 100 or at 80
            while th1>0:
                if print_detail: print(f"Looking for {n.top_temp}({th1}){n.middle_temp}({th2}){n.bottom_temp}")
                node = [
                    x for x in self.nodes[n.time_slice+1]
                    if x.top_temp==n.top_temp
                    and x.middle_temp==n.middle_temp
                    and x.bottom_temp==n.bottom_temp
                    and x.thermocline1==th1
                    and x.thermocline2==th2
                ][0]
                if print_detail: print(f"Energy: {round(node.energy,2)}")
                th1 += -1
                th2 += -1
                candidate_nodes.append(node)
                if next_node_energy >= node.energy:
                    break
            # If the top was at 100 try now 80
            if n.top_temp == 100:
                th1 = self.params.num_layers
                th2 = self.params.num_layers
                while th1>0:
                    if print_detail: print(f"Looking for {n.top_temp-20}({th1}){n.middle_temp-20}({th2}){n.bottom_temp-20}")
                    node = [
                        x for x in self.nodes[n.time_slice+1]
                        if x.top_temp==n.top_temp-20
                        and x.middle_temp==n.middle_temp-20
                        and x.bottom_temp==n.bottom_temp-20
                        and x.thermocline1==th1
                        and x.thermocline2==th2
                    ][0]
                    if print_detail: print(f"Energy: {round(node.energy,2)}")
                    th1 += -1
                    th2 += -1
                    candidate_nodes.append(node)
                    if next_node_energy >= node.energy:
                        break
            closest_node = min([x for x in candidate_nodes], key=lambda x: abs(x.energy-next_node_energy))
            return closest_node

        need_to_break = False
        while True:
            # Moving up step by step until the end of the top layer
            while th1>0:
                if print_detail: print(f"Looking for {n.top_temp}({th1}){n.middle_temp}({th2}){n.bottom_temp}")
                node = [
                    x for x in self.nodes[n.time_slice+1]
                    if x.top_temp==n.top_temp
                    and x.middle_temp==n.middle_temp
                    and x.bottom_temp==n.bottom_temp
                    and x.thermocline1==th1
                    and x.thermocline2==th2
                ][0]
                if print_detail: print(f"Energy: {round(node.energy,2)}")
                th1 += -1
                th2 += -1
                candidate_nodes.append(node)
                need_to_break = True
                if next_node_energy >= node.energy:
                    break
            if need_to_break: break

            # There is no middle layer (cold nodes reached)
            if n.middle_temp == n.bottom_temp or (n.bottom_temp==100 and th1==th2):
                # Go through the top being at 100
                top_temp = 100
                middle_temp = 80
                bottom_temp = 80
                th1 = self.params.num_layers
                th2 = self.params.num_layers
                while th1>0:
                    if print_detail: print(f"Looking for {top_temp}({th1}){middle_temp}({th2}){bottom_temp}")
                    node = [
                        x for x in self.nodes[n.time_slice+1]
                        if x.top_temp==top_temp
                        and x.middle_temp==middle_temp
                        and x.bottom_temp==bottom_temp
                        and x.thermocline1==th1
                        and x.thermocline2==th2
                    ][0]
                    if print_detail: print(f"Energy: {round(node.energy,2)}")
                    th1 += -1
                    th2 += -1
                    candidate_nodes.append(node)
                    if next_node_energy >= node.energy:
                        break
                # Go through the top being at 80
                top_temp = 80
                middle_temp = 60
                bottom_temp = 60
                th1 = self.params.num_layers
                th2 = self.params.num_layers
                while th1>0:
                    if print_detail: print(f"Looking for {top_temp}({th1}){middle_temp}({th2}){bottom_temp}")
                    node = [
                        x for x in self.nodes[n.time_slice+1]
                        if x.top_temp==top_temp
                        and x.middle_temp==middle_temp
                        and x.bottom_temp==bottom_temp
                        and x.thermocline1==th1
                        and x.thermocline2==th2
                    ][0]
                    if print_detail: print(f"Energy: {round(node.energy,2)}")
                    th1 += -1
                    th2 += -1
                    candidate_nodes.append(node)
                    if next_node_energy >= node.energy:
                        break
                closest_node = min([x for x in candidate_nodes], key=lambda x: abs(x.energy-next_node_energy))
                return closest_node

            # Moving up step by step until the end of the middle layer
            top_temp = n.middle_temp
            th1 = th2
            while th1>0:
                if print_detail: print(f"Looking for {top_temp}({th1})-({th2}){n.bottom_temp}")
                node = [
                    x for x in self.nodes[n.time_slice+1]
                    if x.top_temp==top_temp
                    and x.bottom_temp==n.bottom_temp
                    and x.thermocline1==th1
                    and x.thermocline2==th2
                ][0]
                if print_detail: print(f"Energy: {round(node.energy,2)}")
                th1 += -1
                th2 += -1
                candidate_nodes.append(node)
                if next_node_energy >= node.energy:
                    break
            break

        # Find the candidate node which has closest energy to the target
        closest_node = min([x for x in candidate_nodes], key=lambda x: abs(x.energy-next_node_energy))
        return closest_node
        
    def find_closest_node(self, true_n: DNode, print_detail: bool) -> DNode:
        if print_detail: print(f"Looking for closest of {true_n}")

        # Find closest available top, middle and bottom temps
        closest_top_temp = min(self.top_temps, key=lambda x: abs(float(x)-true_n.top_temp))
        closest_middle_temp = min(self.middle_temps, key=lambda x: abs(float(x)-true_n.middle_temp))
        closest_bottom_temp = min(self.bottom_temps, key=lambda x: abs(float(x)-true_n.bottom_temp))

        # Top temperature is impossible to reach
        if true_n.top_temp > max(self.top_temps):
            nodes_with_similar_temps = [
                n for n in self.nodes[true_n.time_slice] if 
                n.top_temp == max(self.top_temps) and
                n.middle_temp==closest_middle_temp and
                n.bottom_temp==closest_bottom_temp and
                n.thermocline2==true_n.thermocline2
            ]
            closest_node = min(nodes_with_similar_temps, key = lambda x: abs(x.energy-true_n.energy))
            return closest_node

        # Both top and middle were rounded above
        if closest_top_temp > true_n.top_temp and closest_middle_temp > true_n.middle_temp and true_n.bottom_temp==100:
            nodes_with_similar_temps = [
                n for n in self.nodes[true_n.time_slice] if 
                n.top_temp<=closest_top_temp and
                n.top_temp>=closest_top_temp-STEP_F and
                n.middle_temp<=closest_middle_temp and
                n.middle_temp>=closest_middle_temp-STEP_F and
                n.bottom_temp==closest_bottom_temp and
                n.thermocline2==true_n.thermocline2
            ]
            closest_node = min(nodes_with_similar_temps, key = lambda x: abs(x.energy-true_n.energy))
            return closest_node
        
        # Both top and middle were rounded below
        if (closest_top_temp < true_n.top_temp and closest_bottom_temp < true_n.middle_temp 
            and true_n.bottom_temp==100):
            nodes_with_similar_temps = [
                n for n in self.nodes[true_n.time_slice] if 
                n.top_temp<=closest_top_temp+STEP_F and
                n.top_temp>=closest_top_temp and
                n.middle_temp<=closest_middle_temp+STEP_F and
                n.middle_temp>=closest_middle_temp and
                n.bottom_temp==closest_bottom_temp and
                n.thermocline2==true_n.thermocline2
            ]
            if print_detail: print(nodes_with_similar_temps)
            closest_node = min(nodes_with_similar_temps, key = lambda x: abs(x.energy-true_n.energy))
            return closest_node

        # Need at least MIN_DIFFERENCE_F between top and middle
        if closest_top_temp == closest_middle_temp or closest_top_temp-5 == closest_middle_temp:
            closest_middle_temp = closest_top_temp-MIN_DIFFERENCE_F if closest_top_temp>100+2*MIN_DIFFERENCE_F else 100
        if closest_top_temp == 120 and closest_middle_temp==100:
            closest_middle_temp = 110
        if true_n.top_temp<=100:
            closest_top_temp = true_n.top_temp
            closest_middle_temp = true_n.top_temp-20
            closest_bottom_temp = true_n.top_temp-20
        if print_detail: print(f"{closest_top_temp},{closest_middle_temp},{closest_bottom_temp}")
        if true_n.thermocline1==true_n.thermocline2:
            nodes_with_similar_temps = [
                n for n in self.nodes[true_n.time_slice] if 
                n.top_temp==closest_top_temp and 
                n.middle_temp==closest_middle_temp and
                n.bottom_temp==closest_bottom_temp and
                n.thermocline1==n.thermocline2 and
                n.thermocline2==true_n.thermocline1
            ]
        else:
            nodes_with_similar_temps = [
                n for n in self.nodes[true_n.time_slice] if 
                n.top_temp==closest_top_temp and 
                n.middle_temp==closest_middle_temp and 
                n.bottom_temp==closest_bottom_temp and
                n.thermocline1==true_n.thermocline1 and
                n.thermocline2==true_n.thermocline2
            ]
        closest_node = min(nodes_with_similar_temps, key = lambda x: abs(x.energy-true_n.energy))
        if len(nodes_with_similar_temps)>1:
            print(true_n)
            print(nodes_with_similar_temps)
            raise Exception("IMPOSSIBLE")
        return closest_node
    
    def solve_dijkstra(self):
        for time_slice in range(self.params.horizon-1, -1, -1):
            for node in self.nodes[time_slice]:
                best_edge = min(self.edges[node], key=lambda e: e.head.pathcost + e.cost)
                node.pathcost = best_edge.head.pathcost + best_edge.cost
                node.next_node = best_edge.head

    def generate_bid(self):
        self.pq_pairs: List[PriceQuantityUnitless] = []
        forecasted_price_usd_mwh = self.params.elec_price_forecast[0] * 10
        # For every possible price
        min_elec_ctskwh, max_elec_ctskwh = -10, 200
        for elec_price_usd_mwh in sorted(list(range(min_elec_ctskwh*10, max_elec_ctskwh*10))+[forecasted_price_usd_mwh]):
            # Update the fake cost of initial node edges with the selected price
            for edge in self.edges[self.initial_node]:
                if edge.cost >= 1e4: # penalized node
                    edge.fake_cost = edge.cost
                else:
                    cop = self.params.COP(oat=self.params.oat_forecast[0], lwt=edge.head.top_temp)
                    edge.fake_cost = edge.hp_heat_out / cop * elec_price_usd_mwh/1000
            # Find the best edge with the given price
            best_edge: DEdge = min(self.edges[self.initial_node], key=lambda e: e.head.pathcost + e.fake_cost)
            if best_edge.hp_heat_out < 0: 
                best_edge_neg = max([e for e in self.edges[self.initial_node] if e.hp_heat_out<0], key=lambda e: e.hp_heat_out)
                best_edge_pos = min([e for e in self.edges[self.initial_node] if e.hp_heat_out>=0], key=lambda e: e.hp_heat_out)
                best_edge = best_edge_pos if (-best_edge_neg.hp_heat_out >= best_edge_pos.hp_heat_out) else best_edge_neg
            # Find the associated quantity
            cop = self.params.COP(oat=self.params.oat_forecast[0], lwt=best_edge.head.top_temp)
            best_quantity_kwh = best_edge.hp_heat_out / cop
            best_quantity_kwh = 0 if best_quantity_kwh<0 else best_quantity_kwh
            if not self.pq_pairs:
                self.pq_pairs.append(
                    PriceQuantityUnitless(
                        PriceTimes1000 = int(elec_price_usd_mwh * 1000),
                        QuantityTimes1000 = int(best_quantity_kwh * 1000))
                )
            else:
                # Record a new pair if at least 0.01 kWh of difference in quantity with the previous one
                if self.pq_pairs[-1].QuantityTimes1000 - int(best_quantity_kwh * 1000) > 10:
                    self.pq_pairs.append(
                        PriceQuantityUnitless(
                            PriceTimes1000 = int(elec_price_usd_mwh * 1000),
                            QuantityTimes1000 = int(best_quantity_kwh * 1000))
                    )
        return self.pq_pairs

    def plot(self, show=True):
        print(f"Time in charge: {round(self.time_spent_in_charge,1)}")
        print(f"Time in discharge: {round(self.time_spent_in_discharge,1)}")
        print(f"Time doing other stuff: {round(self.time_spent_in_total-self.time_spent_in_charge-self.time_spent_in_discharge,1)}")
        print(f"Time in TOTAL: {round(self.time_spent_in_total,1)}")
        # Walk along the shortest path (sp)
        sp_top_temp = []
        sp_middle_temp = []
        sp_bottom_temp = []
        sp_thermocline = []
        sp_thermocline2 = []
        sp_hp_heat_out = []
        sp_stored_energy = []
        node_i = self.initial_node
        the_end = False
        while not the_end:
            if node_i.next_node is None:
                the_end = True
                sp_hp_heat_out.append(edge_i.hp_heat_out)
            else:
                edge_i = [e for e in self.edges[node_i] if e.head==node_i.next_node][0]
                losses = self.params.storage_losses_percent/100 * (node_i.energy-self.bottom_node_energy)
                energy_to_store = edge_i.hp_heat_out-self.params.load_forecast[node_i.time_slice]-losses
                model_energy_to_store = edge_i.head.energy-edge_i.tail.energy
                if model_energy_to_store>energy_to_store:
                    print(f"\n{edge_i}, model thinks {abs(round(model_energy_to_store-energy_to_store,1))} kWh more in store than reality")
                elif model_energy_to_store<energy_to_store:
                    print(f"\n{edge_i}, model thinks {abs(round(model_energy_to_store-energy_to_store,1))} kWh less in store than reality")
                else:
                    print(f"\n{edge_i}, model could not be more accurate!")
                sp_hp_heat_out.append(edge_i.hp_heat_out)
                _ = self.model_accurately(node_i, energy_to_store, print_detail=True)
            sp_top_temp.append(node_i.top_temp)
            sp_bottom_temp.append(node_i.bottom_temp)
            sp_thermocline.append(node_i.thermocline1)
            sp_middle_temp.append(node_i.middle_temp)
            sp_thermocline2.append(node_i.thermocline2)
            sp_stored_energy.append(node_i.energy)
            node_i = node_i.next_node
        sp_soc = [(x-self.bottom_node_energy) / (self.top_node_energy-self.bottom_node_energy) * 100 
                    for x in sp_stored_energy]
        sp_time = list(range(self.params.horizon+1))
        
        # Plot the shortest path
        fig, ax = plt.subplots(2,1, sharex=True, figsize=(10,6))
        start = datetime.fromtimestamp(self.params.start_time, tz=pytz.timezone("America/New_York")).strftime('%Y-%m-%d %H:%M')
        end = (datetime.fromtimestamp(self.params.start_time, tz=pytz.timezone("America/New_York")) 
               + timedelta(hours=self.params.horizon)).strftime('%Y-%m-%d %H:%M')
        fig.suptitle(f'From {start} to {end}\nCost: {round(self.initial_node.pathcost,2)} $', fontsize=10)
        
        # Top plot
        ax[0].step(sp_time, sp_hp_heat_out, where='post', color='tab:blue', alpha=0.6, label='HP')
        ax[0].step(sp_time[:-1], self.params.load_forecast, where='post', color='tab:red', alpha=0.6, label='Load')
        ax[0].legend(loc='upper left')
        ax[0].set_ylabel('Heat [kWh]')
        if max(sp_hp_heat_out)>0:
            ax[0].set_ylim([-0.5, 1.5*max(sp_hp_heat_out)])
        ax2 = ax[0].twinx()
        ax2.step(sp_time[:-1], self.params.elec_price_forecast, where='post', color='gray', alpha=0.6, label='Elec price')
        ax2.legend(loc='upper right')
        ax2.set_ylabel('Electricity price [cts/kWh]')
        m = 0 if min(self.params.elec_price_forecast)>0 else min(self.params.elec_price_forecast)-5
        ax2.set_ylim([m,max(self.params.elec_price_forecast)*1.3])
        
        # Bottom plot
        norm = Normalize(vmin=60, vmax=175+STEP_F)
        cmap = matplotlib.colormaps['Reds']
        tank_top_colors = [cmap(norm(x)) for x in sp_top_temp]
        tank_middle_colors = [cmap(norm(x)) for x in sp_middle_temp]
        tank_bottom_colors = [cmap(norm(x)) for x in sp_bottom_temp]

        # Reversing thermocline positions
        sp_thermocline_reversed1 = [self.params.num_layers - x for x in sp_thermocline]
        sp_thermocline_reversed2 = [self.params.num_layers - x for x in sp_thermocline2]

        # Stacking the temperatures and thermoclines
        bars_top = ax[1].bar(sp_time, 
                             sp_thermocline, 
                             bottom=sp_thermocline_reversed1, 
                             color=tank_top_colors, alpha=0.7, width=0.9) #, align='edge')
        bars_middle = ax[1].bar(sp_time, 
                                [y-x for x,y in zip(sp_thermocline, sp_thermocline2)], 
                                bottom=sp_thermocline_reversed2, 
                                color=tank_middle_colors, alpha=0.7, width=0.9) #, align='edge')
        bars_bottom = ax[1].bar(sp_time, 
                                sp_thermocline_reversed2, 
                                bottom=0, 
                                color=tank_bottom_colors, alpha=0.7, width=0.9) #, align='edge')
        ax[1].set_xlabel('Time [hours]')
        ax[1].set_ylabel('Storage state')
        ax[1].set_ylim([0, self.params.num_layers])
        ax[1].set_yticks([])
        if len(sp_time)>10 and len(sp_time)<50:
            ax[1].set_xticks(list(range(0,len(sp_time)+1,2)))
        for i, bar in enumerate(bars_top):
            height = bar.get_height()
            bar_color = 'white'
            if i < len(self.params.rswt_forecast) and self.params.rswt_forecast[i] <= sp_top_temp[i]:
                bar_color = 'green'
            elif sp_top_temp[i]<100:
                bar_color = 'gray'
            ax[1].text(bar.get_x() + bar.get_width() / 2, bar.get_y() + height / 2, 
                    f'{int(sp_top_temp[i])}', ha='center', va='center', color=bar_color, fontsize=5)
        for i, bar in enumerate(bars_middle):
            height = bar.get_height()
            bar_color = 'white'
            if i < len(self.params.rswt_forecast) and self.params.rswt_forecast[i] <= sp_middle_temp[i]:
                bar_color = 'green'
            elif sp_middle_temp[i]<100:
                bar_color = 'gray'
            if height>1:
                ax[1].text(bar.get_x() + bar.get_width() / 2, bar.get_y() + height / 2, 
                        f'{int(sp_middle_temp[i])}', ha='center', va='center', color=bar_color, fontsize=5)
        for i, bar in enumerate(bars_bottom):
            height = bar.get_height()
            bar_color = 'white'
            if i < len(self.params.rswt_forecast) and self.params.rswt_forecast[i] <= sp_bottom_temp[i]:
                bar_color = 'green'
            elif sp_bottom_temp[i]<100:
                bar_color = 'gray'
            ax[1].text(bar.get_x() + bar.get_width() / 2, bar.get_y() + height / 2, 
                    f'{int(sp_bottom_temp[i])}', ha='center', va='center', color=bar_color, fontsize=5)
        ax3 = ax[1].twinx()
        ax3.plot(sp_time, sp_soc, color='black', alpha=0.4, label='SoC')
        ax3.set_ylabel('State of charge [%]')
        ax3.set_ylim([-1,101])

        # Color bar
        boundaries = sorted(list(range(60,175,5)), reverse=False)
        colors = [plt.cm.Reds(i/(len(boundaries)-1)) for i in range(len(boundaries))]
        cmap = ListedColormap(colors)
        norm = BoundaryNorm(boundaries, cmap.N, clip=True)
        sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
        cbar = plt.colorbar(sm, ax=ax, orientation='horizontal', fraction=0.06, pad=0.15, alpha=0.7)
        cbar.set_ticks(sorted(list(range(60,175,5)), reverse=False))
        cbar.set_label('Temperature [F]')
        
        plt.savefig('plot.png', dpi=130)
        if show:
            plt.show()
        plt.close()

    def export_to_excel(self):        

        # Add the parameters to a seperate sheet
        parameters = self.params.config.to_dict()
        parameters_df = pd.DataFrame(list(parameters.items()), columns=['Variable', 'Value'])

        # Add the PQ pairs to a seperate sheet and plot the curve
        pq_pairs = self.generate_bid()
        prices = [x.PriceTimes1000 for x in pq_pairs]
        quantities = [x.QuantityTimes1000/1000 for x in pq_pairs]
        pqpairs_df = pd.DataFrame({'price':[x/1000 for x in prices], 'quantity':quantities})
        # To plot quantities on x-axis and prices on y-axis
        ps, qs = [], []
        index_p = 0
        expected_price_usd_mwh = self.params.elec_price_forecast[0] * 10
        for p in sorted(list(range(min(prices), max(prices)+1)) + [expected_price_usd_mwh*1000]):
            ps.append(p/1000)
            if index_p+1 < len(prices) and p >= prices[index_p+1]:
                index_p += 1
            if p == expected_price_usd_mwh*1000:
                interesection = (quantities[index_p], expected_price_usd_mwh)
            qs.append(quantities[index_p])
        plt.plot(qs, ps, label='demand (bid)')
        prices = [x.PriceTimes1000/1000 for x in pq_pairs]
        plt.scatter(quantities, prices)
        plt.plot([min(quantities)-1, max(quantities)+1],[expected_price_usd_mwh]*2, label="supply (expected market price)")
        plt.scatter(interesection[0], interesection[1])
        plt.text(interesection[0]+0.25, interesection[1]+15, f'({round(interesection[0],3)}, {round(interesection[1],1)})', fontsize=10, color='tab:orange')
        plt.xticks(quantities)
        if min([abs(x-expected_price_usd_mwh) for x in prices]) < 5:
            plt.yticks(prices)
        else:
            plt.yticks(prices + [expected_price_usd_mwh])
        plt.ylabel("Price [USD/MWh]")
        plt.xlabel("Quantity [kWh]")
        plt.grid(alpha=0.3)
        plt.legend()
        plt.tight_layout()
        plt.savefig('plot_pq.png', dpi=130)
        plt.close()

        # Write to Excel
        start = datetime.fromtimestamp(self.params.start_time, tz=pytz.timezone("America/New_York")).strftime('%Y-%m-%d %H:%M')
        # os.makedirs('results', exist_ok=True)
        # file_path = os.path.join('results', f'result_{start}.xlsx')
        file_path = 'result.xlsx'
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:

            self.plot(show=True)
            plot_sheet = writer.book.create_sheet(title='Plot')
            plot_sheet.add_image(Image('plot.png'), 'A1')

            parameters_df.to_excel(writer, index=False, sheet_name='Parameters')

            plot2_sheet = writer.book.create_sheet(title='PQ pairs')
            pqpairs_df.to_excel(writer, index=False, sheet_name='PQ pairs')
            plot2_sheet.add_image(Image('plot_pq.png'), 'C1')

        os.remove('plot.png')        
        os.remove('plot_pq.png')


if __name__ == '__main__':
    import json
    from models import MessageSql

    with open('messages.json', 'r') as file:
        data = json.load(file)
        message = MessageSql(
            message_id=data["MessageId"],
            from_alias=data["FromAlias"],
            message_type_name=data["MessageTypeName"],
            message_persisted_ms=data["MessagePersistedMs"],
            payload=data["Payload"],
            message_created_ms=data.get("MessageCreatedMs")
        )
    flo_params = FloParamsHouse0(**message.payload)
    flo_params.NumLayers = NUM_LAYERS
    flo_params.InitialThermocline = int(NUM_LAYERS/2)
    
    st = time.time()
    g = DGraph(flo_params)
    g.solve_dijkstra()
    g.generate_bid()
    print(f"Built graph and solved Dijkstra in {round(time.time()-st,1)} seconds")
    g.plot()
    # g.export_to_excel()

    # Compare with original FLO
    # st = time.time()
    # from flo import DGraph
    # g = DGraph(flo_params)
    # g.solve_dijkstra()
    # print(f"Built graph and solved Dijkstra in {round(time.time()-st,1)} seconds")
    # g.plot()