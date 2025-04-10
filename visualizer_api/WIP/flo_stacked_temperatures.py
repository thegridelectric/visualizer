from datetime import datetime, timedelta
import numpy as np
from typing import Dict, List, Tuple
import os
import pytz
import pandas as pd
import matplotlib
import matplotlib.pyplot as plt
from matplotlib.dates import DateFormatter
from matplotlib.colors import Normalize, ListedColormap, BoundaryNorm
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.drawing.image import Image
from named_types import PriceQuantityUnitless, FloParamsHouse0
from typing import Optional

HOUSES_WITH_ONLY_2_EDGES = ['oak']

def to_kelvin(t):
    return (t-32)*5/9 + 273.15

def to_celcius(t):
    return (t-32)*5/9

class DParams():
    def __init__(self, config: FloParamsHouse0) -> None:
        self.config = config
        self.start_time = config.StartUnixS
        self.horizon = config.HorizonHours
        self.num_layers = config.NumLayers
        self.storage_volume = config.StorageVolumeGallons
        self.max_hp_elec_in = config.HpMaxElecKw
        self.min_hp_elec_in = config.HpMinElecKw
        self.initial_top_temp = config.InitialTopTempF
        self.initial_bottom_temp = config.InitialBottomTempF
        self.initial_thermocline = config.InitialThermocline
        self.storage_losses_percent = config.StorageLossesPercent
        self.reg_forecast = [x/10 for x in config.RegPriceForecast[:self.horizon]]
        self.dist_forecast = [x/10 for x in config.DistPriceForecast[:self.horizon]]
        self.lmp_forecast = [x/10 for x in config.LmpForecast[:self.horizon]]
        self.elec_price_forecast = [rp+dp+lmp for rp,dp,lmp in zip(self.reg_forecast, self.dist_forecast, self.lmp_forecast)]
        self.oat_forecast = config.OatForecastF[:self.horizon]
        self.ws_forecast = config.WindSpeedForecastMph[:self.horizon]
        self.alpha = config.AlphaTimes10/10
        self.beta = config.BetaTimes100/100
        self.gamma = config.GammaEx6/1e6
        self.no_power_rswt = -self.alpha/self.beta
        self.intermediate_power = config.IntermediatePowerKw
        self.intermediate_rswt = config.IntermediateRswtF
        self.dd_power = config.DdPowerKw
        self.dd_rswt = config.DdRswtF
        self.dd_delta_t = config.DdDeltaTF
        self.hp_is_off = config.HpIsOff
        self.hp_turn_on_minutes = config.HpTurnOnMinutes
        self.quadratic_coefficients = self.get_quadratic_coeffs()
        self.temperature_stack = self.get_available_top_temps()
        self.load_forecast = [self.required_heating_power(oat,ws) for oat,ws in zip(self.oat_forecast,self.ws_forecast)]
        self.rswt_forecast = [self.required_swt(x) for x in self.load_forecast]
        # Modify load forecast to include energy available in the buffer
        available_buffer = config.BufferAvailableKwh
        i = 0
        while available_buffer > 0 and i < len(self.load_forecast):
            load_backup = self.load_forecast[i]
            self.load_forecast[i] = self.load_forecast[i] - min(available_buffer, self.load_forecast[i])
            available_buffer = available_buffer - min(available_buffer, load_backup)
            i += 1
        # Modify load forecast to include energy available in the house (zones above thermostat)
        available_house = config.HouseAvailableKwh
        i = 0
        if available_house < 0:
            self.load_forecast[0] += -available_house
        else:
            while available_house > 0 and i < len(self.load_forecast):
                load_backup = self.load_forecast[i]
                self.load_forecast[i] = self.load_forecast[i] - min(available_house, self.load_forecast[i])
                available_house = available_house - min(available_house, load_backup)
                i += 1
        self.check_hp_sizing()
        # TODO: add to config
        self.min_cop = 1
        self.max_cop = 3
        self.soft_constraint: bool = True
        # First time step can be shorter than an hour
        if datetime.fromtimestamp(self.start_time).minute > 0:
            self.fraction_of_hour_remaining: float = datetime.fromtimestamp(self.start_time).minute / 60
        else:
            self.fraction_of_hour_remaining: float = 1
        self.load_forecast[0] = self.load_forecast[0]*self.fraction_of_hour_remaining
        # Find the first top temperature above RSWT for each hour
        self.rswt_plus = {}
        for rswt in self.rswt_forecast:
            self.rswt_plus[rswt] = self.first_top_temp_above_rswt(rswt)
        
    def check_hp_sizing(self):
        max_load_elec = max(self.load_forecast) / self.COP(min(self.oat_forecast), max(self.rswt_forecast))
        if max_load_elec > self.max_hp_elec_in:
            error_text = f"\nThe current parameters indicate that on the coldest hour of the forecast ({min(self.oat_forecast)} F):"
            error_text += f"\n- The heating requirement is {round(max(self.load_forecast),2)} kW"
            error_text += f"\n- The COP is {round(self.COP(min(self.oat_forecast), max(self.rswt_forecast)),2)}"
            error_text += f"\n=> Need a HP that can reach {round(max_load_elec,2)} kW electrical power"
            error_text += f"\n=> The given HP is undersized ({self.max_hp_elec_in} kW electrical power)"
            print(error_text)
        
    def COP(self, oat, lwt=None):
        if oat < self.config.CopMinOatF: 
            return self.config.CopMin
        else:
            return self.config.CopIntercept + self.config.CopOatCoeff * oat

    def required_heating_power(self, oat, ws):
        r = self.alpha + self.beta*oat + self.gamma*ws
        return r if r>0 else 0

    def delivered_heating_power(self, swt):
        a, b, c = self.quadratic_coefficients
        d = a*swt**2 + b*swt + c
        return d if d>0 else 0

    def required_swt(self, rhp):
        a, b, c = self.quadratic_coefficients
        c2 = c - rhp
        return (-b + (b**2-4*a*c2)**0.5)/(2*a)

    def delta_T(self, swt):
        d = self.dd_delta_t/self.dd_power * self.delivered_heating_power(swt)
        d = 0 if swt<self.no_power_rswt else d
        return d if d>0 else 0
    
    def delta_T_inverse(self, rwt: float) -> float:
        a, b, c = self.quadratic_coefficients
        aa = -self.dd_delta_t/self.dd_power * a
        bb = 1-self.dd_delta_t/self.dd_power * b
        cc = -self.dd_delta_t/self.dd_power * c - rwt
        if bb**2-4*aa*cc < 0 or (-bb + (bb**2-4*aa*cc)**0.5)/(2*aa) - rwt > 30:
            return 30
        dt = (-bb + (bb**2-4*aa*cc)**0.5)/(2*aa) - rwt
        if dt<=1:
            return 1
        return dt
    
    def get_quadratic_coeffs(self):
        x_rswt = np.array([self.no_power_rswt, self.intermediate_rswt, self.dd_rswt])
        y_hpower = np.array([0, self.intermediate_power, self.dd_power])
        A = np.vstack([x_rswt**2, x_rswt, np.ones_like(x_rswt)]).T
        return [float(x) for x in np.linalg.solve(A, y_hpower)] 
    
    def get_available_top_temps(self) -> Tuple[Dict, Dict]:
        MIN_BOTTOM_TEMP, MAX_TOP_TEMP = 80, 175

        if self.initial_top_temp < MIN_BOTTOM_TEMP:
            self.initial_top_temp = MIN_BOTTOM_TEMP

        if self.initial_bottom_temp < self.initial_top_temp - self.delta_T(self.initial_top_temp):
            self.initial_bottom_temp = round(self.initial_top_temp - self.delta_T(self.initial_top_temp))

        if self.initial_bottom_temp < MIN_BOTTOM_TEMP:
            self.initial_bottom_temp = MIN_BOTTOM_TEMP

        if self.initial_bottom_temp == self.initial_top_temp:
            self.initial_bottom_temp += -5

        self.max_thermocline = self.num_layers
        if self.initial_top_temp > MAX_TOP_TEMP-5:
            self.max_thermocline = self.initial_thermocline

        available_temps = []
        height_top = self.initial_thermocline
        height_bottom = self.num_layers - self.initial_thermocline

        # Add temperatures above initial tank
        t = self.initial_top_temp
        b = self.initial_bottom_temp
        while t < MAX_TOP_TEMP or b < MAX_TOP_TEMP:
            if t > MAX_TOP_TEMP:
                available_temps.append((b, height_bottom))
                b = round(b + self.delta_T_inverse(b))
            else:
                available_temps.append((b, height_bottom))
                available_temps.append((t, height_top))
                if t==round(t + self.delta_T_inverse(t)) or b==round(b + self.delta_T_inverse(b)):
                    break
                t = round(t + self.delta_T_inverse(t))
                b = round(b + self.delta_T_inverse(b))

        # Add temperatures below initial tank
        t = round(self.initial_top_temp - self.delta_T(self.initial_top_temp))
        b = round(self.initial_bottom_temp - self.delta_T(self.initial_bottom_temp))
        while b > MIN_BOTTOM_TEMP or t > MIN_BOTTOM_TEMP:
            if b < MIN_BOTTOM_TEMP:
                available_temps = [(t, height_top)] + available_temps
                t = round(t - self.delta_T(t))
            else:
                available_temps = [(t, height_top)] + available_temps
                available_temps = [(b, height_bottom)] + available_temps
                if t==round(t - self.delta_T(t)) or b==round(b - self.delta_T(b)):
                    break
                t = round(t - self.delta_T(t))
                b = round(b - self.delta_T(b))

        self.available_top_temps = [x[0] for x in available_temps]
        if self.available_top_temps != sorted(self.available_top_temps):
            for i in range(1, len(available_temps)):
                available_temps[i] = (max(available_temps[i][0], available_temps[i-1][0]), available_temps[i][1])

        for _ in range(10):
            if sorted(list(set([x[0] for x in available_temps]))) == sorted([x[0] for x in available_temps]):
                break
            available_temps_no_duplicates = []
            skip_next_i = False
            for i in range(len(available_temps)):
                if i<len(available_temps)-1 and available_temps[i][0] == available_temps[i+1][0]:
                    available_temps_no_duplicates.append((available_temps[i][0], min(
                        available_temps[i][1]+available_temps[i+1][1], self.num_layers)))
                    skip_next_i = True
                elif not skip_next_i:
                    available_temps_no_duplicates.append(available_temps[i])
                else:
                    skip_next_i = False
            available_temps = available_temps_no_duplicates.copy()

        if max([x[0] for x in available_temps]) < MAX_TOP_TEMP-5:
            available_temps.append((MAX_TOP_TEMP-5, self.num_layers))

        if self.max_thermocline == self.num_layers and available_temps[-1][1] < self.num_layers:
            available_temps[-1] = (available_temps[-1][0], self.num_layers)

        self.available_top_temps = [x[0] for x in available_temps]
        if self.available_top_temps != sorted(self.available_top_temps):
            print("ERROR sorted is not the same")

        # heights = [x[1] for x in available_temps]
        # fig, ax = plt.subplots(figsize=(8, 6))
        # cmap = matplotlib.colormaps['Reds']
        # norm = plt.Normalize(min(self.available_top_temps)-20, max(self.available_top_temps)+20)
        # bottom = 0
        # for i in range(len(available_temps)):
        #     color = cmap(norm(self.available_top_temps[i]))
        #     ax.bar(0, heights[i], bottom=bottom, color=color, width=1)
        #     ax.text(0, bottom + heights[i]/2, str(self.available_top_temps[i]), ha='center', va='center', fontsize=10, color='white')
        #     if i < len(available_temps)-1:
        #         bottom += heights[i]
        # ax.set_xticks([])
        # ax.set_xlim([-2,2])
        # plt.title(self.initial_top_temp)
        # plt.tight_layout()
        # plt.show()

        self.energy_between_nodes = {}
        m_layer = self.storage_volume*3.785 / self.num_layers
        for i in range(1,len(self.available_top_temps)):
            temp_drop_f = self.available_top_temps[i] - self.available_top_temps[i-1]
            self.energy_between_nodes[self.available_top_temps[i]] = round(m_layer * 4.187/3600 * temp_drop_f*5/9,3)

        return available_temps

    def first_top_temp_above_rswt(self, rswt):
        for x in sorted(self.available_top_temps):
            if x > rswt:
                return x

class DNode():
    def __init__(self, time_slice:int, top_temp:float, thermocline1:float, parameters:DParams, hinge_node=None):
        self.params = parameters
        # Position in graph
        self.time_slice = time_slice
        self.top_temp = top_temp
        self.thermocline1 = thermocline1
        if not hinge_node:
            temperatures = [x[0] for x in self.params.temperature_stack]
            heights = [x[1] for x in self.params.temperature_stack]
            toptemp_idx = temperatures.index(top_temp)
            height_first_two_layers = thermocline1 + heights[toptemp_idx-1]
            if height_first_two_layers >= self.params.num_layers or toptemp_idx < 2:
                self.middle_temp = None
                self.bottom_temp = temperatures[toptemp_idx-1]
                self.thermocline2 = None
            else:
                self.middle_temp = temperatures[toptemp_idx-1]
                self.bottom_temp = temperatures[toptemp_idx-2]
                self.thermocline2 = height_first_two_layers
            # Dijkstra's algorithm
            self.pathcost = 0 if time_slice==parameters.horizon else 1e9
            self.next_node = None
            # Absolute energy level
            self.energy = self.get_energy()
            self.index = None
        else:
            self.middle_temp = hinge_node['middle_temp']
            self.thermocline2 = hinge_node['thermocline2']
            self.bottom_temp = hinge_node['bottom_temp']
            self.pathcost = hinge_node['pathcost']
            self.energy = self.get_energy()

    def __repr__(self):
        if self.thermocline2 is not None:
            return f"{self.top_temp}({self.thermocline1}){self.middle_temp}({self.thermocline2}){self.bottom_temp}"
            # return f"Node[top:{self.top_temp}, thermocline1:{self.thermocline1}, middle:{self.middle_temp}, thermocline2:{self.thermocline2}, bottom:{self.bottom_temp}]"
        else:
            return f"{self.top_temp}({self.thermocline1}){self.bottom_temp}"
            # return f"Node[top:{self.top_temp}, thermocline1:{self.thermocline1}, bottom:{self.bottom_temp}]"

    def get_energy(self):
        m_layer_kg = self.params.storage_volume*3.785 / self.params.num_layers
        if self.middle_temp is not None:
            kWh_top = (self.thermocline1-0.5)*m_layer_kg * 4.187/3600 * to_kelvin(self.top_temp)
            kWh_midlle = (self.thermocline2-self.thermocline1)*m_layer_kg * 4.187/3600 * to_kelvin(self.middle_temp)
            kWh_bottom = (self.params.num_layers-self.thermocline2+0.5)*m_layer_kg * 4.187/3600 * to_kelvin(self.bottom_temp)
        else:        
            kWh_top = (self.thermocline1-0.5)*m_layer_kg * 4.187/3600 * to_kelvin(self.top_temp)
            kWh_midlle = 0
            kWh_bottom = (self.params.num_layers-self.thermocline1+0.5)*m_layer_kg * 4.187/3600 * to_kelvin(self.bottom_temp)
        return kWh_top + kWh_midlle + kWh_bottom


class DEdge():
    def __init__(self, tail:DNode, head:DNode, cost:float, hp_heat_out:float, rswt_minus_edge_elec:Optional[float]=None):
        self.tail: DNode = tail
        self.head: DNode = head
        self.cost = cost
        self.hp_heat_out = hp_heat_out
        self.rswt_minus_edge_elec = rswt_minus_edge_elec
        self.fake_cost: Optional[float] = None

    def __repr__(self):
        return f"Edge: {self.tail} --cost:{round(self.cost,3)}, heat:{round(self.hp_heat_out,2)}--> {self.head}"


class DGraph():
    def __init__(self, config: FloParamsHouse0):
        self.params = DParams(config)
        self.nodes: Dict[int, List[DNode]] = {}
        self.edges: Dict[DNode, List[DEdge]] = {}
        self.create_nodes()
        self.create_edges()

    def create_nodes(self):
        self.initial_node = DNode(0, self.params.initial_top_temp, self.params.initial_thermocline, self.params)
        for time_slice in range(self.params.horizon+1):
            self.nodes[time_slice] = [self.initial_node] if time_slice==0 else []
            if self.params.max_thermocline < self.params.num_layers:
                self.nodes[time_slice].extend(
                    DNode(time_slice, top_temp, thermocline, self.params)
                    for top_temp in self.params.available_top_temps[1:-1]
                    for thermocline in range(1,self.params.num_layers+1)
                    if thermocline <= self.params.temperature_stack[self.params.available_top_temps.index(top_temp)][1]
                    and (time_slice, top_temp, thermocline) != (0, self.params.initial_top_temp, self.params.initial_thermocline)
                )
                self.nodes[time_slice].extend(
                    DNode(time_slice, self.params.available_top_temps[-1], thermocline, self.params)
                    for thermocline in range(1,self.params.max_thermocline+1)
                    if thermocline <= self.params.temperature_stack[-1][1]
                    and (time_slice, self.params.available_top_temps[-1], thermocline) != (0, self.params.initial_top_temp, self.params.initial_thermocline)
                )
            else:
                self.nodes[time_slice].extend(
                    DNode(time_slice, top_temp, thermocline, self.params)
                    for top_temp in self.params.available_top_temps[1:]
                    for thermocline in range(1,self.params.num_layers+1)
                    if thermocline <= self.params.temperature_stack[self.params.available_top_temps.index(top_temp)][1]
                    and (time_slice, top_temp, thermocline) != (0, self.params.initial_top_temp, self.params.initial_thermocline)
                )

    def create_edges(self):
        
        self.bottom_node = DNode(0, 
                                 self.params.available_top_temps[1],
                                 self.params.num_layers - self.params.temperature_stack[self.params.available_top_temps.index(self.params.available_top_temps[0])][1],
                                 self.params)
        self.top_node = DNode(0, 
                              self.params.available_top_temps[-1], 
                              self.params.temperature_stack[self.params.available_top_temps.index(self.params.available_top_temps[-1])][1], 
                              self.params)
        
        only_2_edges = False
        for house_alias in HOUSES_WITH_ONLY_2_EDGES:
            if house_alias in self.params.config.GNodeAlias:
                only_2_edges = True
                print(f"Only two edges for {house_alias}")
        
        for h in range(self.params.horizon):
            
            for node_now in self.nodes[h]:
                self.edges[node_now] = []

                # The losses might be lower than energy between two nodes
                losses = self.params.storage_losses_percent/100 * (node_now.energy-self.bottom_node.energy)
                if self.params.load_forecast[h]==0 and losses>0 and losses<self.params.energy_between_nodes[node_now.top_temp]:
                    losses = self.params.energy_between_nodes[node_now.top_temp] + 1/1e9

                # If the current top temperature is the first one available above RSWT
                # If it exists, add an edge from the current node that drains the storage further than RSWT
                if not only_2_edges:
                    RSWT_plus = self.params.rswt_plus[self.params.rswt_forecast[h]]
                    if node_now.top_temp == RSWT_plus and h < self.params.horizon-1:
                        self.add_rswt_minus_edge(node_now, h, losses)

                for node_next in self.nodes[h+1]:

                    store_heat_in = node_next.energy - node_now.energy
                    hp_heat_out = store_heat_in + self.params.load_forecast[h] + losses
                    
                    # Adjust the max elec the HP can use in the first time step
                    # (Duration of time step + turn-on effects)
                    max_hp_elec_in = self.params.max_hp_elec_in
                    if h==0:
                        max_hp_elec_in = max_hp_elec_in * self.params.fraction_of_hour_remaining
                        max_hp_elec_in = (((1-self.params.hp_turn_on_minutes/60) if self.params.hp_is_off else 1) * max_hp_elec_in)
                    
                    # This condition reduces the amount of times we need to compute the COP
                    if (hp_heat_out/self.params.max_cop <= max_hp_elec_in and
                        hp_heat_out/self.params.min_cop >= self.params.min_hp_elec_in):
                    
                        cop = self.params.COP(oat=self.params.oat_forecast[h], lwt=node_next.top_temp)

                        if (hp_heat_out/cop <= max_hp_elec_in and 
                            hp_heat_out/cop >= self.params.min_hp_elec_in):

                            cost = self.params.elec_price_forecast[h]/100 * hp_heat_out/cop

                            # If some of the load is satisfied by the storage
                            # Then it must satisfy the SWT requirement
                            if store_heat_in < 0:
                                if ((hp_heat_out < self.params.load_forecast[h] and 
                                     self.params.load_forecast[h] > 0)
                                     and
                                    (node_now.top_temp < self.params.rswt_forecast[h] or 
                                     node_next.top_temp < self.params.rswt_forecast[h])):
                                    if self.params.soft_constraint and not [x for x in self.edges[node_now] if x.head==node_next]:
                                        cost += 1e5
                                    else:
                                        continue
                            
                            self.edges[node_now].append(DEdge(node_now, node_next, cost, hp_heat_out))
                
                if self.edges[node_now] and only_2_edges:
                    min_hp_out_edge = min(self.edges[node_now], key=lambda e: e.hp_heat_out)
                    max_hp_out_edge = max(self.edges[node_now], key=lambda e: e.hp_heat_out)
                    self.edges[node_now] = [min_hp_out_edge]
                    if max_hp_out_edge.hp_heat_out > 10:
                        self.edges[node_now].append(max_hp_out_edge)
                    else:
                        self.edges[node_now][0].hp_heat_out = 0

                if not self.edges[node_now]:
                    print(f"No edge from node {node_now}, adding edge with penalty")
                    cop = self.params.COP(oat=self.params.oat_forecast[h], lwt=node_next.top_temp)
                    hp_heat_out = max_hp_elec_in * cop
                    node_next = [n for n in self.nodes[h+1] if n.top_temp==node_now.top_temp and n.thermocline1==node_now.thermocline1][0]
                    self.edges[node_now].append(DEdge(node_now, node_next, 1e5, hp_heat_out))
                    print(DEdge(node_now, node_next, 1e5, hp_heat_out))

    def add_rswt_minus_edge(self, node: DNode, time_slice, Q_losses):
        # In these calculations the load is both the heating requirement and the losses
        Q_load = self.params.load_forecast[time_slice] + Q_losses
        # Find the heat stored in the water that is hotter than RSWT
        Q_plus = 0
        m_layer_kg = self.params.storage_volume*3.785 / self.params.num_layers
        RSWT_minus = node.top_temp
        if node.top_temp > self.params.rswt_forecast[time_slice]:
            m_plus = (node.thermocline1-0.5) * m_layer_kg
            Q_plus = m_plus * 4.187/3600 * self.params.delta_T(node.top_temp)*5/9
            if node.middle_temp is not None:
                RSWT_minus = node.middle_temp
                RSWT_minus_minus = node.bottom_temp
            else:
                RSWT_minus = node.bottom_temp
                RSWT_minus_minus = None
        # The hot part of the storage alone can not provide the load and the losses
        if Q_plus <= Q_load:
            RSWT = self.params.rswt_forecast[time_slice]
            RSWT_min_DT = RSWT - self.params.delta_T(RSWT)
            m_chc = Q_load / (4.187/3600 * (to_kelvin(RSWT)-to_kelvin(RSWT_min_DT)))
            m_minus_max = (1 - Q_plus/Q_load) * m_chc
            if m_minus_max > self.params.storage_volume*3.785:
                print("m_minus_max > total storage mass!")
                m_minus_max = self.params.storage_volume*3.785
            if node.middle_temp is not None and RSWT_minus != node.top_temp:
                m_minus = m_layer_kg * (node.thermocline2 - node.thermocline1)
                if m_minus > m_minus_max:
                    m_minus = m_minus_max
                m_minus_minus = m_minus_max - m_minus
                Q_minus_max = (
                    m_minus * 4.187/3600 * (self.params.delta_T(RSWT_minus)*5/9)
                    + m_minus_minus * 4.187/3600 * (self.params.delta_T(RSWT_minus_minus)*5/9)
                    )
            else:
                Q_minus_max = m_minus_max * 4.187/3600 * (self.params.delta_T(RSWT_minus)*5/9)
            Q_missing = Q_load - Q_plus - Q_minus_max
            if Q_missing < 0:
                print(f"Isn't this impossible? Q_load - Q_plus - Q_minus_max = {round(Q_missing,2)} kWh")
            Q_missing = 0 if Q_missing < 0 else Q_missing
            if Q_missing > 0 and not self.params.soft_constraint:
                return
            # Find the next node that would be the closest to matching the energy
            next_node_energy = node.energy - (Q_plus + Q_minus_max)
            next_nodes = [x for x in self.nodes[time_slice+1] if x.energy <= node.energy and x.energy >= next_node_energy]
            next_node = sorted(next_nodes, key=lambda x: x.energy)[0]
            # Penalty is slightly more expensive than the cost of producing Q_missing in the next hour
            cop = self.params.COP(oat=self.params.oat_forecast[time_slice+1], lwt=next_node.top_temp)
            penalty = (Q_missing+1)/cop * self.params.elec_price_forecast[time_slice+1]/100
            self.edges[node].append(DEdge(node, next_node, penalty, 0, rswt_minus_edge_elec=(Q_missing+1)/cop))

    def solve_dijkstra(self):
        for time_slice in range(self.params.horizon-1, -1, -1):
            for node in self.nodes[time_slice]:
                best_edge = min(self.edges[node], key=lambda e: e.head.pathcost + e.cost)
                if best_edge.hp_heat_out < 0: 
                    best_edge_neg = max([e for e in self.edges[node] if e.hp_heat_out<0], key=lambda e: e.hp_heat_out)
                    best_edge_pos = min([e for e in self.edges[node] if e.hp_heat_out>=0], key=lambda e: e.hp_heat_out)
                    best_edge = best_edge_pos if (-best_edge_neg.hp_heat_out >= best_edge_pos.hp_heat_out) else best_edge_neg
                node.pathcost = best_edge.head.pathcost + best_edge.cost
                node.next_node = best_edge.head
    
    def generate_bid(self):
        self.pq_pairs: List[PriceQuantityUnitless] = []
        forecasted_price_usd_mwh = self.params.elec_price_forecast[0] * 10
        # For every possible price
        min_elec_ctskwh, max_elec_ctskwh = -10, 200
        for elec_price_usd_mwh in sorted(list(range(min_elec_ctskwh*10, max_elec_ctskwh*10))+[forecasted_price_usd_mwh]):
            # Update the fake cost of initial node edges with the selected price
            for edge in self.edges[self.initial_node]:
                if edge.cost >= 1e4: # penalized node
                    edge.fake_cost = edge.cost
                elif edge.rswt_minus_edge_elec is not None: # penalized node
                    edge.fake_cost = edge.rswt_minus_edge_elec * elec_price_usd_mwh/1000
                else:
                    cop = self.params.COP(oat=self.params.oat_forecast[0], lwt=edge.head.top_temp)
                    edge.fake_cost = edge.hp_heat_out / cop * elec_price_usd_mwh/1000
            # Find the best edge with the given price
            best_edge: DEdge = min(self.edges[self.initial_node], key=lambda e: e.head.pathcost + e.fake_cost)
            if best_edge.hp_heat_out < 0: 
                best_edge_neg = max([e for e in self.edges[self.initial_node] if e.hp_heat_out<0], key=lambda e: e.hp_heat_out)
                best_edge_pos = min([e for e in self.edges[self.initial_node] if e.hp_heat_out>=0], key=lambda e: e.hp_heat_out)
                best_edge = best_edge_pos if (-best_edge_neg.hp_heat_out >= best_edge_pos.hp_heat_out) else best_edge_neg
            # Find the associated quantity
            cop = self.params.COP(oat=self.params.oat_forecast[0], lwt=best_edge.head.top_temp)
            best_quantity_kwh = best_edge.hp_heat_out / cop
            best_quantity_kwh = 0 if best_quantity_kwh<0 else best_quantity_kwh
            if not self.pq_pairs:
                self.pq_pairs.append(
                    PriceQuantityUnitless(
                        PriceTimes1000 = int(elec_price_usd_mwh * 1000),
                        QuantityTimes1000 = int(best_quantity_kwh * 1000))
                )
            else:
                # Record a new pair if at least 0.01 kWh of difference in quantity with the previous one
                if self.pq_pairs[-1].QuantityTimes1000 - int(best_quantity_kwh * 1000) > 10:
                    self.pq_pairs.append(
                        PriceQuantityUnitless(
                            PriceTimes1000 = int(elec_price_usd_mwh * 1000),
                            QuantityTimes1000 = int(best_quantity_kwh * 1000))
                    )
        return self.pq_pairs
    
    def quick_plot(self, show=True):
        # Walk along the shortest path (sp)
        sp_top_temp = []
        sp_middle_temp = []
        sp_bottom_temp = []
        sp_thermocline = []
        sp_thermocline2 = []
        sp_hp_heat_out = []
        sp_stored_energy = []
        node_i = self.initial_node
        the_end = False
        while not the_end:
            if node_i.next_node is None:
                the_end = True
                sp_hp_heat_out.append(edge_i.hp_heat_out)
            else:
                edge_i = [e for e in self.edges[node_i] if e.head==node_i.next_node][0]
                sp_hp_heat_out.append(edge_i.hp_heat_out)
            sp_top_temp.append(node_i.top_temp)
            sp_bottom_temp.append(node_i.bottom_temp)
            sp_thermocline.append(node_i.thermocline1)
            if node_i.middle_temp is not None:
                sp_middle_temp.append(node_i.middle_temp)
                sp_thermocline2.append(node_i.thermocline2)
            else:
                sp_middle_temp.append(node_i.bottom_temp)
                sp_thermocline2.append(node_i.thermocline1)
            sp_stored_energy.append(node_i.energy)
            node_i = node_i.next_node
        sp_soc = [(x-self.bottom_node.energy) / (self.top_node.energy-self.bottom_node.energy) * 100 
                    for x in sp_stored_energy]
        sp_time = list(range(self.params.horizon+1))
        start_time = datetime.fromtimestamp(self.params.start_time, tz=pytz.timezone("America/New_York"))
        sp_time = [(start_time+timedelta(hours=x)) for x in range(len(sp_time))]
        
        # Plot the shortest path
        fig, ax = plt.subplots(2,1, figsize=(12,6), gridspec_kw={'height_ratios':[8,6]})
        plt.subplots_adjust(hspace=0.3) 
        start = datetime.fromtimestamp(self.params.start_time, tz=pytz.timezone("America/New_York")).strftime('%Y-%m-%d %H:%M')
        
        # Top plot
        plot_hours = 12
        ax[0].step(sp_time[:plot_hours], sp_hp_heat_out[:plot_hours], where='post', color='tab:red', alpha=0.6, label='HP')
        ax[0].step(sp_time[:plot_hours], self.params.load_forecast[:plot_hours], where='post', color='black', linestyle='dashed', alpha=0.4, label='Load')
        ax[0].legend(loc='upper left')
        ax[0].set_title(f'{start}', fontsize=10)
        ax[0].set_ylabel('Heat [kWh]')
        ax[0].set_ylim([-0.5, 1.5*max(sp_hp_heat_out)])
        ax2 = ax[0].twinx()
        ax2.step(sp_time[:plot_hours], self.params.lmp_forecast[:plot_hours], where='post', color='tab:green', alpha=0.8, label='LMP')
        
        ax2.set_ylabel('Electricity price [cts/kWh]')
        yticks = list(set([int(x) for x in self.params.lmp_forecast[:plot_hours]]))
        yticks = sorted(yticks+[x+0.5 for x in yticks])
        if len(ax2.get_yticks())>=6 and len(yticks)<=6:
            ax2.set_yticks(yticks)
        ax[0].set_xticks([x for x in sp_time][:plot_hours])
        ax[0].set_xticklabels([f'{x.hour}:00' for x in sp_time][:plot_hours])

        # Bottom plot
        ax[1].plot(sp_time[:plot_hours], sp_soc[:plot_hours], color='black', alpha=0.4, label='SoC')
        ax[1].set_ylabel('Energy in the store [kWh]')
        ax[1].set_ylim([max(-1,min(sp_soc[:plot_hours])-10),101])
        ax[1].set_yticks([])
        ax[1].set_xticks([x for x in sp_time][:plot_hours])
        ax[1].set_xticklabels([f'{x.hour}:00' for x in sp_time][:plot_hours])

        done_mornings = {}
        done_afternoons = {}
        not_labeled = True
        for i, x in enumerate(sp_time[:plot_hours]):
            if x.hour in [7,8,9,10,11,16,17,18,19] and x.weekday() not in [5,6]:
                if x.hour in [7,8,9,10,11] and x.date() not in done_mornings:
                    end_index = i+5-(x.hour-7) if i==0 else min(i+5, plot_hours-1)
                    done_mornings[x.date()] = True
                    if not_labeled:
                        ax2.axvspan(sp_time[i], sp_time[end_index], color='tab:green', alpha=0.05, label='Onpeak')
                        not_labeled = False
                    else:
                        ax2.axvspan(sp_time[i], sp_time[end_index], color='tab:green', alpha=0.05)
                elif x.hour in [16,17,18,19] and x.date() not in done_afternoons:
                    end_index = i+4-(x.hour-16) if i==0 else min(i+4, plot_hours-1)
                    done_afternoons[x.date()] = True
                    if not_labeled:
                        ax2.axvspan(sp_time[i], sp_time[end_index], color='tab:green', alpha=0.05, label='Onpeak')
                        not_labeled = False
                    else:
                        ax2.axvspan(sp_time[i], sp_time[end_index], color='tab:green', alpha=0.05)
        ax2.legend(loc='upper right')

        # plt.tight_layout()
        plt.savefig('plot_quick.png', dpi=100, bbox_inches='tight')
        if show:
            plt.show()
        plt.close()
    
    def plot(self, show=True):
        # Walk along the shortest path (sp)
        sp_top_temp = []
        sp_middle_temp = []
        sp_bottom_temp = []
        sp_thermocline = []
        sp_thermocline2 = []
        sp_hp_heat_out = []
        sp_stored_energy = []
        node_i = self.initial_node
        the_end = False
        while not the_end:
            if node_i.next_node is None:
                the_end = True
                sp_hp_heat_out.append(edge_i.hp_heat_out)
            else:
                edge_i = [e for e in self.edges[node_i] if e.head==node_i.next_node][0]
                sp_hp_heat_out.append(edge_i.hp_heat_out)
            sp_top_temp.append(node_i.top_temp)
            sp_bottom_temp.append(node_i.bottom_temp)
            sp_thermocline.append(node_i.thermocline1)
            if node_i.middle_temp is not None:
                sp_middle_temp.append(node_i.middle_temp)
                sp_thermocline2.append(node_i.thermocline2)
            else:
                sp_middle_temp.append(node_i.bottom_temp)
                sp_thermocline2.append(node_i.thermocline1)
            sp_stored_energy.append(node_i.energy)
            node_i = node_i.next_node
        sp_soc = [(x-self.bottom_node.energy) / (self.top_node.energy-self.bottom_node.energy) * 100 
                    for x in sp_stored_energy]
        sp_time = list(range(self.params.horizon+1))
        
        # Plot the shortest path
        fig, ax = plt.subplots(2,1, sharex=True, figsize=(10,6))
        start = datetime.fromtimestamp(self.params.start_time, tz=pytz.timezone("America/New_York")).strftime('%Y-%m-%d %H:%M')
        end = (datetime.fromtimestamp(self.params.start_time, tz=pytz.timezone("America/New_York")) + timedelta(hours=self.params.horizon)).strftime('%Y-%m-%d %H:%M')
        fig.suptitle(f'From {start} to {end}\nCost: {round(self.initial_node.pathcost,2)} $', fontsize=10)
        
        # Top plot
        ax[0].step(sp_time, sp_hp_heat_out, where='post', color='tab:blue', alpha=0.6, label='HP')
        ax[0].step(sp_time[:-1], self.params.load_forecast, where='post', color='tab:red', alpha=0.6, label='Load')
        ax[0].legend(loc='upper left')
        ax[0].set_ylabel('Heat [kWh]')
        ax[0].set_ylim([-0.5, 1.5*max(sp_hp_heat_out)])
        ax2 = ax[0].twinx()
        ax2.step(sp_time[:-1], self.params.elec_price_forecast, where='post', color='gray', alpha=0.6, label='Elec price')
        ax2.legend(loc='upper right')
        ax2.set_ylabel('Electricity price [cts/kWh]')
        m = 0 if min(self.params.elec_price_forecast)>0 else min(self.params.elec_price_forecast)-5
        ax2.set_ylim([m,max(self.params.elec_price_forecast)*1.3])
        
        # Bottom plot
        norm = Normalize(vmin=self.params.available_top_temps[0], vmax=self.params.available_top_temps[-1])
        cmap = matplotlib.colormaps['Reds']
        tank_top_colors = [cmap(norm(x)) for x in sp_top_temp]
        tank_middle_colors = [cmap(norm(x)) for x in sp_middle_temp]
        tank_bottom_colors = [cmap(norm(x)) for x in sp_bottom_temp]

        # Reversing thermocline positions
        sp_thermocline_reversed1 = [self.params.num_layers - x for x in sp_thermocline]
        sp_thermocline_reversed2 = [self.params.num_layers - x for x in sp_thermocline2]

        # Stacking the temperatures and thermoclines
        bars_top = ax[1].bar(sp_time, sp_thermocline, bottom=sp_thermocline_reversed1, color=tank_top_colors, alpha=0.7, width=0.9, align='edge')
        bars_middle = ax[1].bar(sp_time, [y-x for x,y in zip(sp_thermocline, sp_thermocline2)], bottom=sp_thermocline_reversed2, color=tank_middle_colors, alpha=0.7, width=0.9, align='edge')
        bars_bottom = ax[1].bar(sp_time, sp_thermocline_reversed2, bottom=0, color=tank_bottom_colors, alpha=0.7, width=0.9, align='edge')
        ax[1].set_xlabel('Time [hours]')
        ax[1].set_ylabel('Storage state')
        ax[1].set_ylim([0, self.params.num_layers])
        ax[1].set_yticks([])
        if len(sp_time)>10 and len(sp_time)<50:
            ax[1].set_xticks(list(range(0,len(sp_time)+1,2)))
        for i, bar in enumerate(bars_top):
            height = bar.get_height()
            bar_color = 'white'
            if i < len(self.params.rswt_forecast) and self.params.rswt_forecast[i] <= sp_top_temp[i]:
                bar_color = 'green'
            elif height<3:
                bar_color='black'
            ax[1].text(bar.get_x() + bar.get_width() / 2, bar.get_y() + height / 2, 
                    f'{int(sp_top_temp[i])}', ha='center', va='center', color=bar_color, fontsize=5)
        for i, bar in enumerate(bars_middle):
            height = bar.get_height()
            bar_color = 'white'
            if i < len(self.params.rswt_forecast) and self.params.rswt_forecast[i] <= sp_middle_temp[i]:
                bar_color = 'green'
            if height>1:
                ax[1].text(bar.get_x() + bar.get_width() / 2, bar.get_y() + height / 2, 
                        f'{int(sp_middle_temp[i])}', ha='center', va='center', color=bar_color, fontsize=5)
        for i, bar in enumerate(bars_bottom):
            height = bar.get_height()
            bar_color = 'white'
            if i < len(self.params.rswt_forecast) and self.params.rswt_forecast[i] <= sp_bottom_temp[i]:
                bar_color = 'green'
            ax[1].text(bar.get_x() + bar.get_width() / 2, bar.get_y() + height / 2, 
                    f'{int(sp_bottom_temp[i])}', ha='center', va='center', color=bar_color, fontsize=5)
        ax3 = ax[1].twinx()
        ax3.plot(sp_time, sp_soc, color='black', alpha=0.4, label='SoC')
        ax3.set_ylabel('State of charge [%]')
        ax3.set_ylim([-1,101])

        # Color bar
        boundaries = self.params.available_top_temps
        colors = [plt.cm.Reds(i/(len(boundaries)-1)) for i in range(len(boundaries))]
        cmap = ListedColormap(colors)
        norm = BoundaryNorm(boundaries, cmap.N, clip=True)
        sm = plt.cm.ScalarMappable(cmap=cmap, norm=norm)
        cbar = plt.colorbar(sm, ax=ax, orientation='horizontal', fraction=0.06, pad=0.15, alpha=0.7)
        cbar.set_ticks(self.params.available_top_temps)
        cbar.set_label('Temperature [F]')
        
        plt.savefig('plot.png', dpi=130)
        if show:
            plt.show()
        plt.close()

    def export_to_excel(self):        
        # Sort nodes by energy and assign an index
        for time_slice in range(self.params.horizon+1):
            self.nodes_by_energy = sorted(self.nodes[time_slice], key=lambda x: (x.energy, x.top_temp), reverse=True)
            for n in self.nodes[time_slice]:
                n.index = self.nodes_by_energy.index(n)+1

        # Along the shortest path
        electricitiy_used, heat_delivered = [], []
        node_i = self.initial_node
        while node_i.next_node is not None:
            losses = self.params.storage_losses_percent/100 * (node_i.energy-self.bottom_node.energy)
            if self.params.load_forecast[node_i.time_slice]==0 and losses>0 and losses<self.params.energy_between_nodes[node_i.top_temp]:
                losses = self.params.energy_between_nodes[node_i.top_temp] + 1/1e9
            store_heat_in = node_i.next_node.energy - node_i.energy
            hp_heat_out = store_heat_in + self.params.load_forecast[node_i.time_slice] + losses
            cop = self.params.COP(oat=self.params.oat_forecast[node_i.time_slice], lwt=node_i.next_node.top_temp)
            heat_delivered.append(hp_heat_out)
            electricitiy_used.append(hp_heat_out/cop)
            node_i = node_i.next_node
        
        # First dataframe: the Dijkstra graph
        dijkstra_pathcosts = {}
        dijkstra_pathcosts['Model'] = [repr(x) for x in self.nodes_by_energy]
        dijkstra_pathcosts['Energy (relative)'] = [round(x.energy-self.bottom_node.energy,2) for x in self.nodes_by_energy]
        dijkstra_pathcosts['Index'] = list(range(1,len(self.nodes_by_energy)+1))
        dijkstra_nextnodes = dijkstra_pathcosts.copy()
        for h in range(self.params.horizon):
            dijkstra_pathcosts[h] = [round(x.pathcost,2) for x in sorted(self.nodes[h], key=lambda x: x.index)]
            dijkstra_nextnodes[h] = [x.next_node.index for x in sorted(self.nodes[h], key=lambda x: x.index)]
        dijkstra_pathcosts[self.params.horizon] = [0 for x in self.nodes[self.params.horizon]]
        dijkstra_nextnodes[self.params.horizon] = [np.nan for x in self.nodes[self.params.horizon]]
        dijkstra_pathcosts_df = pd.DataFrame(dijkstra_pathcosts)
        dijkstra_nextnodes_df = pd.DataFrame(dijkstra_nextnodes)
        
        # Second dataframe: the forecasts
        start_time = datetime.fromtimestamp(self.params.start_time, tz=pytz.timezone("America/New_York"))
        forecast_df = pd.DataFrame({'Forecast':['0'], 'Unit':['0'], **{h: [0.0] for h in range(self.params.horizon)}})
        forecast_df.loc[0] = ['Hour'] + [start_time.strftime("%d/%m/%Y")] + [(start_time + timedelta(hours=x)).hour for x in range(self.params.horizon)]
        forecast_df.loc[1] = ['Price - total'] + ['cts/kWh'] + self.params.elec_price_forecast
        forecast_df.loc[2] = ['Price - distribution'] + ['cts/kWh'] + self.params.dist_forecast
        forecast_df.loc[3] = ['Price - LMP'] + ['cts/kWh'] + self.params.lmp_forecast
        forecast_df.loc[4] = ['Heating load'] + ['kW'] + [round(x,2) for x in self.params.load_forecast]
        forecast_df.loc[5] = ['OAT'] + ['F'] + [round(x,2) for x in self.params.oat_forecast]
        forecast_df.loc[6] = ['Required SWT'] + ['F'] + [round(x) for x in self.params.rswt_forecast]
        
        # Third dataframe: the shortest path
        shortestpath_df = pd.DataFrame({'Shortest path':['0'], 'Unit':['0'], **{h: [0.0] for h in range(self.params.horizon+1)}})
        shortestpath_df.loc[0] = ['Electricity used'] + ['kWh'] + [round(x,3) for x in electricitiy_used] + [0]
        shortestpath_df.loc[1] = ['Heat delivered'] + ['kWh'] + [round(x,3) for x in heat_delivered] + [0]
        shortestpath_df.loc[2] = ['Cost - total'] + ['cts'] + [round(x*y,2) for x,y in zip(electricitiy_used, self.params.elec_price_forecast)] + [0]
        shortestpath_df.loc[3] = ['Cost - distribution'] + ['cts'] + [round(x*y,2) for x,y in zip(electricitiy_used, self.params.dist_forecast)] + [0]
        shortestpath_df.loc[4] = ['Cost - LMP'] + ['cts'] + [round(x*y,2) for x,y in zip(electricitiy_used, self.params.lmp_forecast)] + [0]
        
        # Fourth dataframe: the results
        total_usd = round(self.initial_node.pathcost,2)
        total_elec = round(sum(electricitiy_used),2)
        total_heat = round(sum(heat_delivered),2)
        next_index = self.initial_node.next_node.index
        results = ['Cost ($)', total_usd, 'Electricity (kWh)', total_elec, 'Heat (kWh)', total_heat, 'Next step index', next_index]
        results_df = pd.DataFrame({'RESULTS':results})
        
        # Highlight shortest path
        highlight_positions = []
        node_i = self.initial_node
        while node_i.next_node is not None:
            highlight_positions.append((node_i.index+len(forecast_df)+len(shortestpath_df)+2, 3+node_i.time_slice))
            node_i = node_i.next_node
        highlight_positions.append((node_i.index+len(forecast_df)+len(shortestpath_df)+2, 3+node_i.time_slice))
        
        # Add the parameters to a seperate sheet
        parameters = self.params.config.to_dict()
        parameters_df = pd.DataFrame(list(parameters.items()), columns=['Variable', 'Value'])

        # Add the PQ pairs to a seperate sheet and plot the curve
        pq_pairs = self.generate_bid()
        prices = [x.PriceTimes1000 for x in pq_pairs]
        quantities = [x.QuantityTimes1000/1000 for x in pq_pairs]
        pqpairs_df = pd.DataFrame({'price':[x/1000 for x in prices], 'quantity':quantities})
        # To plot quantities on x-axis and prices on y-axis
        ps, qs = [], []
        index_p = 0
        expected_price_usd_mwh = self.params.elec_price_forecast[0] * 10
        for p in sorted(list(range(min(prices), max(prices)+1)) + [expected_price_usd_mwh*1000]):
            ps.append(p/1000)
            if index_p+1 < len(prices) and p >= prices[index_p+1]:
                index_p += 1
            if p == expected_price_usd_mwh*1000:
                interesection = (quantities[index_p], expected_price_usd_mwh)
            qs.append(quantities[index_p])
        plt.plot(qs, ps, label='demand (bid)')
        prices = [x.PriceTimes1000/1000 for x in pq_pairs]
        plt.scatter(quantities, prices)
        plt.plot([min(quantities)-1, max(quantities)+1],[expected_price_usd_mwh]*2, label="supply (expected market price)")
        plt.scatter(interesection[0], interesection[1])
        plt.text(interesection[0]+0.25, interesection[1]+15, f'({round(interesection[0],3)}, {round(interesection[1],1)})', fontsize=10, color='tab:orange')
        plt.xticks(quantities)
        if min([abs(x-expected_price_usd_mwh) for x in prices]) < 5:
            plt.yticks(prices)
        else:
            plt.yticks(prices + [expected_price_usd_mwh])
        plt.ylabel("Price [USD/MWh]")
        plt.xlabel("Quantity [kWh]")
        plt.grid(alpha=0.3)
        plt.legend()
        plt.tight_layout()
        plt.savefig('plot_pq.png', dpi=130)
        plt.close()

        # Write to Excel
        start = datetime.fromtimestamp(self.params.start_time, tz=pytz.timezone("America/New_York")).strftime('%Y-%m-%d %H:%M')
        # os.makedirs('results', exist_ok=True)
        # file_path = os.path.join('results', f'result_{start}.xlsx')
        file_path = 'result.xlsx'
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:

            # Add summary plot
            self.quick_plot(show=False)
            quick_plot_sheet = writer.book.create_sheet(title='Quick look')
            quick_plot_sheet.add_image(Image('plot_quick.png'), 'A1')

            results_df.to_excel(writer, index=False, sheet_name='Pathcost')
            results_df.to_excel(writer, index=False, sheet_name='Next node')
            forecast_df.to_excel(writer, index=False, startcol=1, sheet_name='Pathcost')
            forecast_df.to_excel(writer, index=False, startcol=1, sheet_name='Next node')
            shortestpath_df.to_excel(writer, index=False, startcol=1, startrow=len(forecast_df)+1, sheet_name='Pathcost')
            shortestpath_df.to_excel(writer, index=False, startcol=1, startrow=len(forecast_df)+1, sheet_name='Next node')
            dijkstra_pathcosts_df.to_excel(writer, index=False, startrow=len(forecast_df)+len(shortestpath_df)+2, sheet_name='Pathcost')
            dijkstra_nextnodes_df.to_excel(writer, index=False, startrow=len(forecast_df)+len(shortestpath_df)+2, sheet_name='Next node')
            parameters_df.to_excel(writer, index=False, sheet_name='Parameters')

            # Add plot in a seperate sheet
            self.plot(show=False)
            plot_sheet = writer.book.create_sheet(title='Plot')
            plot_sheet.add_image(Image('plot.png'), 'A1')

            # Add plot in a seperate sheet
            plot2_sheet = writer.book.create_sheet(title='PQ pairs')
            pqpairs_df.to_excel(writer, index=False, sheet_name='PQ pairs')
            plot2_sheet.add_image(Image('plot_pq.png'), 'C1')

            # Layout
            pathcost_sheet = writer.sheets['Pathcost']
            nextnode_sheet = writer.sheets['Next node']
            parameters_sheet = writer.sheets['Parameters']
            for row in pathcost_sheet['A1:A10']:
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
            for row in nextnode_sheet['A1:A10']:
                for cell in row:
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold=True)
            for row in parameters_sheet[f'B1:B{len(parameters_df)+1}']:
                for cell in row:
                    cell.alignment = Alignment(horizontal='right')
            pathcost_sheet.column_dimensions['A'].width = 17.5
            pathcost_sheet.column_dimensions['B'].width = 15
            pathcost_sheet.column_dimensions['C'].width = 15
            nextnode_sheet.column_dimensions['A'].width = 17.5
            nextnode_sheet.column_dimensions['B'].width = 15
            nextnode_sheet.column_dimensions['C'].width = 15
            parameters_sheet.column_dimensions['A'].width = 40
            parameters_sheet.column_dimensions['B'].width = 70
            pathcost_sheet.freeze_panes = 'D16'
            nextnode_sheet.freeze_panes = 'D16'

            # Highlight shortest path
            highlight_fill = PatternFill(start_color='72ba93', end_color='72ba93', fill_type='solid')
            for row in range(len(forecast_df)+len(shortestpath_df)+2):
                pathcost_sheet.cell(row=row+1, column=1).fill = highlight_fill
                nextnode_sheet.cell(row=row+1, column=1).fill = highlight_fill
            for row, col in highlight_positions:
                pathcost_sheet.cell(row=row+1, column=col+1).fill = highlight_fill
                nextnode_sheet.cell(row=row+1, column=col+1).fill = highlight_fill

        os.remove('plot.png')        
        os.remove('plot_quick.png')        
        os.remove('plot_pq.png')